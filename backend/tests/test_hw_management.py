"""Hardware Management: depreciation engine, registers, budget summary, workbooks.

The engine tests replay the golden cells lifted from
`HW_purchasing_working_document_V5.xlsx`; the API tests drive the same numbers
through `/api/hw`. Renewal-risk dates are built relative to `date.today()` so the
suite does not rot, and every project/catalog row a test needs is created by that
test — nothing here assumes a seeded database.
"""

import io
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pytest
import xlsxwriter
from alembic import command
from alembic.config import Config
from sqlalchemy import create_engine, inspect

from app.config import HW_LEASING_MONTHS, HW_PURCHASE_TYPES
from app.database import Base, run_migrations
from app.routers.hardware import XLSX_MEDIA_TYPE
from app.services import hw_depreciation as dep
from app.services import hw_excel

HW_TABLES = {"hw_projects", "hw_assets", "hw_licenses", "hw_budget_adjustments"}

# purchase, end, cost, kind, year, expected — every row of the spec's table.
GOLDEN_CASES = [
    ("2025-07-02", "2028-07-02", 7157.35, "Leasing", 2025, 1192.8916666666669),
    ("2025-07-02", "2028-07-02", 7157.35, "Leasing", 2026, 2385.7833333333338),
    ("2025-07-02", "2028-07-02", 7157.35, "Leasing", 2028, 1391.7069444444446),
    ("2025-09-01", "2026-02-07", 564.57, "Leasing", 2025, 62.730000000000004),
    ("2025-09-01", "2026-02-07", 564.57, "Leasing", 2026, 31.365000000000002),
    ("2023-08-03", "2026-08-03", 12550.89, "Leasing", 2023, 1743.1791666666666),
    ("2023-08-03", "2026-08-03", 12550.89, "Leasing", 2026, 2789.0866666666666),
    ("2023-12-20", "2026-12-20", 4689.71, "Leasing", 2023, 130.2697222222222),
    ("2025-07-02", "2026-07-09", 9877.0, "Purchase", 2025, 9877.0),
    ("2025-07-02", "2026-07-09", 9877.0, "Purchase", 2026, 0.0),
]

EMPTY_ASSET = {
    "asset_tag": "", "company": "", "name": "", "serial": "", "model": "",
    "category": "", "status": "", "supplier": "", "purchase_date": None,
    "purchase_cost": 0.0, "order_number": "", "eol_date": None,
    "assigned_employee": "", "sw_license": "", "purchased_by": "",
    "purchase_type": "Not Purchased", "catalog_item_id": None,
}

EMPTY_LICENSE = {
    "license_tag": "", "company": "", "name": "", "product_key": "",
    "expiration_date": None, "licensed_to_email": "", "category": "",
    "supplier": "", "manufacturer": "", "quantity": 1, "purchase_date": None,
    "termination_date": None, "depreciation": "Not Purchased", "maintained": False,
    "purchase_cost": 0.0, "purchase_order_number": "", "notes": "",
    "catalog_item_id": None,
}


def asset(**overrides):
    return {**EMPTY_ASSET, **overrides}


def hw_license(**overrides):
    return {**EMPTY_LICENSE, **overrides}


def make_project(client, name, **overrides):
    body = {
        "name": name, "company": "Vehiclevo", "description": "",
        "budget_assets": 0.0, "budget_licenses": 0.0, "portal_reference": "",
        **overrides,
    }
    resp = client.post("/api/hw/projects", json=body)
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


def build_workbook(sheets):
    """An xlsx in memory: `sheets` is [(title, headers, rows), ...]."""
    buffer = io.BytesIO()
    workbook = xlsxwriter.Workbook(buffer, {"in_memory": True})
    for title, headers, rows in sheets:
        worksheet = workbook.add_worksheet(title)
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)
        for index, row in enumerate(rows, start=1):
            for col, value in enumerate(row):
                worksheet.write(index, col, value)
    workbook.close()
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Depreciation engine
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("purchase,end,cost,kind,year,expected", GOLDEN_CASES)
def test_year_cost_matches_workbook(purchase, end, cost, kind, year, expected):
    assert dep.year_cost(
        year, kind, date.fromisoformat(purchase), date.fromisoformat(end), cost,
    ) == pytest.approx(expected, abs=1e-6)


def test_complete_months_edges():
    # Day-of-month rollback: the 31st to the 30th is not a whole second month
    assert dep.complete_months(date(2026, 1, 31), date(2026, 3, 30)) == 1
    assert dep.complete_months(date(2026, 1, 31), date(2026, 3, 31)) == 2
    # Same month, and the same day, are zero whole months
    assert dep.complete_months(date(2026, 5, 1), date(2026, 5, 31)) == 0
    assert dep.complete_months(date(2026, 5, 10), date(2026, 5, 10)) == 0
    # Year boundary
    assert dep.complete_months(date(2025, 11, 15), date(2026, 2, 15)) == 3
    assert dep.complete_months(date(2025, 12, 31), date(2026, 1, 1)) == 0
    assert dep.complete_months(date(2025, 1, 1), date(2026, 1, 1)) == 12


def test_year_cost_edge_cases():
    purchase, end = date(2026, 1, 1), date(2028, 12, 31)
    # Leasing needs both ends of the term
    assert dep.year_cost(2026, "Leasing", None, end, 3600.0) == 0.0
    assert dep.year_cost(2026, "Leasing", purchase, None, 3600.0) == 0.0
    # Years outside the term contribute nothing
    assert dep.year_cost(2025, "Leasing", purchase, end, 3600.0) == 0.0
    assert dep.year_cost(2029, "Leasing", purchase, end, 3600.0) == 0.0
    # Neither planned nor unpurchased rows hit the actual budget
    assert dep.year_cost(2026, "Planned Purchase", purchase, end, 3600.0) == 0.0
    assert dep.year_cost(2026, "Not Purchased", purchase, end, 3600.0) == 0.0
    assert dep.year_cost(2026, "", purchase, end, 3600.0) == 0.0
    # The sheet compares the type case-insensitively
    assert dep.year_cost(2026, " leasing ", purchase, end, 3600.0) == 1200.0
    assert dep.year_cost(2026, "purchase", purchase, end, 3600.0) == 3600.0
    # A 36-month lease amortises exactly once over HW_LEASING_MONTHS
    total = sum(dep.year_cost(y, "Leasing", purchase, end, 3600.0)
                for y in (2026, 2027, 2028))
    assert total == pytest.approx(3600.0)
    assert HW_LEASING_MONTHS == 36


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

def test_hw_migration_round_trip(tmp_path):
    """Down and back up, which is the only path that runs the real create_table.

    The baseline revision builds a fresh database with `create_all`, so the
    0005 guard short-circuits there; downgrading first proves the hand-written
    DDL still produces the schema the ORM maps.
    """
    fresh_engine = create_engine(f"sqlite:///{tmp_path / 'hw.db'}")
    config = Config(str(Path(__file__).resolve().parents[1] / "alembic.ini"))
    try:
        run_migrations(fresh_engine)
        assert set(inspect(fresh_engine).get_table_names()) >= HW_TABLES

        with fresh_engine.begin() as connection:
            config.attributes["connection"] = connection
            command.downgrade(config, "20260819_0004")
        assert not (HW_TABLES & set(inspect(fresh_engine).get_table_names()))

        with fresh_engine.begin() as connection:
            config.attributes["connection"] = connection
            command.upgrade(config, "head")

        inspector = inspect(fresh_engine)
        assert set(inspector.get_table_names()) >= HW_TABLES
        for table in sorted(HW_TABLES):
            mapped = {column.name for column in Base.metadata.tables[table].columns}
            assert {column["name"] for column in inspector.get_columns(table)} == mapped
    finally:
        fresh_engine.dispose()


# ---------------------------------------------------------------------------
# Vocabularies and project CRUD
# ---------------------------------------------------------------------------

def test_hw_meta(client):
    resp = client.get("/api/hw/meta")
    assert resp.status_code == 200
    data = resp.json()
    assert data["purchase_types"] == HW_PURCHASE_TYPES
    assert "In Stock" in data["asset_statuses"]
    assert "Lauterbach debugger" in data["asset_categories"]
    assert "Floating License" in data["license_categories"]
    assert data["leasing_months"] == 36


def test_hw_project_crud(client):
    resp = client.post("/api/hw/projects", json={
        "name": "CRUD Project", "company": "Vehiclevo",
        "description": "Register under test",
        "budget_assets": 1000.0, "budget_licenses": 500.0,
        "portal_reference": "PORTAL-1",
    })
    assert resp.status_code == 201, resp.text
    created = resp.json()
    project_id = created["id"]
    assert created["name"] == "CRUD Project"
    assert created["budget_assets"] == 1000.0
    assert created["portal_reference"] == "PORTAL-1"

    resp = client.get(f"/api/hw/projects/{project_id}")
    assert resp.status_code == 200
    assert resp.json()["description"] == "Register under test"

    resp = client.put(f"/api/hw/projects/{project_id}", json={
        "name": "CRUD Project", "company": "Vehiclevo GmbH", "description": "",
        "budget_assets": 2000.0, "budget_licenses": 500.0, "portal_reference": "",
    })
    assert resp.status_code == 200
    assert resp.json()["company"] == "Vehiclevo GmbH"
    assert resp.json()["budget_assets"] == 2000.0

    rollups = client.get("/api/hw/projects").json()
    row = next(p for p in rollups if p["id"] == project_id)
    assert row["asset_count"] == 0
    assert row["budget_total"] == 2500.0
    assert row["remaining"] == 2500.0

    assert client.delete(f"/api/hw/projects/{project_id}").status_code == 204
    assert client.get(f"/api/hw/projects/{project_id}").status_code == 404


def test_hw_project_404s(client):
    missing = 987654
    for path in (
        f"/api/hw/projects/{missing}",
        f"/api/hw/projects/{missing}/summary",
        f"/api/hw/projects/{missing}/assets",
        f"/api/hw/projects/{missing}/licenses",
        f"/api/hw/projects/{missing}/adjustments",
    ):
        resp = client.get(path)
        assert resp.status_code == 404, path
        assert resp.json()["detail"] == "Hardware project not found"

    resp = client.put(f"/api/hw/projects/{missing}", json={
        "name": "Nope", "company": "", "description": "",
        "budget_assets": 0.0, "budget_licenses": 0.0, "portal_reference": "",
    })
    assert resp.status_code == 404
    assert client.delete(f"/api/hw/projects/{missing}").status_code == 404

    resp = client.post(f"/api/hw/projects/{missing}/assets", json=asset(name="X"))
    assert resp.status_code == 404

    # A project still needs a name
    resp = client.post("/api/hw/projects", json={
        "name": "", "company": "", "description": "",
        "budget_assets": 0.0, "budget_licenses": 0.0, "portal_reference": "",
    })
    assert resp.status_code == 422


# ---------------------------------------------------------------------------
# Registers
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def assets_project(client):
    return make_project(client, "Assets Register", budget_assets=20000.0)


def test_asset_register(client, assets_project):
    # 36-month lease over three calendar years: 3600 / 36 * 12 per year
    resp = client.post(f"/api/hw/projects/{assets_project}/assets", json=asset(
        asset_tag="A-1", name="Lauterbach PowerDebug", category="Lauterbach debugger",
        status="In Stock", purchase_date="2026-01-01", eol_date="2028-12-31",
        purchase_cost=3600.0, purchase_type="Leasing",
    ))
    assert resp.status_code == 201, resp.text
    leased = resp.json()
    assert leased["hw_project_id"] == assets_project
    assert leased["per_year"] == {"2026": 1200.0, "2027": 1200.0, "2028": 1200.0}
    assert leased["total"] == 3600.0

    resp = client.post(f"/api/hw/projects/{assets_project}/assets", json=asset(
        asset_tag="A-2", name="Vector VN1630", category="Vector Box",
        status="Labeled", purchase_date="2026-05-10", purchase_cost=4800.0,
        purchase_type="Purchase",
    ))
    assert resp.status_code == 201
    bought = resp.json()
    assert bought["per_year"] == {"2026": 4800.0, "2027": 0.0, "2028": 0.0}
    assert bought["total"] == 4800.0

    rows = client.get(f"/api/hw/projects/{assets_project}/assets").json()
    assert [row["asset_tag"] for row in rows] == ["A-1", "A-2"]

    # Single-row edit recomputes the year columns
    resp = client.put(f"/api/hw/assets/{leased['id']}", json=asset(
        asset_tag="A-1", name="Lauterbach PowerDebug", category="Lauterbach debugger",
        status="In Stock", purchase_date="2026-01-01", eol_date="2028-12-31",
        purchase_cost=7200.0, purchase_type="Leasing",
    ))
    assert resp.status_code == 200
    assert resp.json()["per_year"] == {"2026": 2400.0, "2027": 2400.0, "2028": 2400.0}

    assert client.delete(f"/api/hw/assets/{bought['id']}").status_code == 204
    assert len(client.get(f"/api/hw/projects/{assets_project}/assets").json()) == 1

    resp = client.put("/api/hw/assets/987654", json=asset(name="X"))
    assert resp.status_code == 404
    assert resp.json()["detail"] == "Hardware asset not found"
    assert client.delete("/api/hw/assets/987654").status_code == 404


def test_asset_bulk_replace(client, assets_project):
    """The grid save posts the whole register; the posted list becomes the truth."""
    before = client.get(f"/api/hw/projects/{assets_project}/assets").json()
    assert before  # the single-row test left a lease behind

    resp = client.put(f"/api/hw/projects/{assets_project}/assets", json={"items": [
        asset(asset_tag="B-1", name="Bench PC", category="PC", status="In Stock",
              purchase_date="2026-01-01", eol_date="2028-12-31",
              purchase_cost=3600.0, purchase_type="Leasing"),
        asset(asset_tag="B-2", name="Power supply", category="Power supply",
              status="Return", purchase_date="2027-02-02", purchase_cost=250.0,
              purchase_type="Purchase"),
    ]})
    assert resp.status_code == 200, resp.text
    rows = resp.json()
    assert [row["asset_tag"] for row in rows] == ["B-1", "B-2"]
    assert rows[0]["per_year"] == {"2026": 1200.0, "2027": 1200.0, "2028": 1200.0}
    assert rows[1]["per_year"] == {"2026": 0.0, "2027": 250.0, "2028": 0.0}
    # The replaced rows are gone, not merged
    assert {row["id"] for row in rows}.isdisjoint({row["id"] for row in before})
    assert len(client.get(f"/api/hw/projects/{assets_project}/assets").json()) == 2

    # An empty list clears the register
    resp = client.put(f"/api/hw/projects/{assets_project}/assets", json={"items": []})
    assert resp.status_code == 200
    assert resp.json() == []


def test_asset_catalog_link(client, assets_project):
    """A catalog link is validated by hand — SQLite does not enforce the FK."""
    catalog = client.post("/api/hardware-catalog", json={
        "name": "HW Mgmt Test Probe", "aspice": "SWE.5", "billing": "once",
        "unit_cost": 4200.0, "supplier_name": "Lauterbach",
    })
    assert catalog.status_code == 201, catalog.text
    catalog_id = catalog.json()["id"]

    resp = client.post(f"/api/hw/projects/{assets_project}/assets", json=asset(
        name="Linked Probe", purchase_cost=4200.0, catalog_item_id=catalog_id,
    ))
    assert resp.status_code == 201
    assert resp.json()["catalog_item_id"] == catalog_id

    resp = client.post(f"/api/hw/projects/{assets_project}/assets", json=asset(
        name="Dangling", catalog_item_id=987654,
    ))
    assert resp.status_code == 404
    assert resp.json()["detail"] == "Catalog item not found: 987654"

    client.put(f"/api/hw/projects/{assets_project}/assets", json={"items": []})


@pytest.fixture(scope="module")
def licenses_project(client):
    return make_project(client, "Licenses Register", budget_licenses=30000.0)


def test_license_register(client, licenses_project):
    resp = client.post(f"/api/hw/projects/{licenses_project}/licenses", json=hw_license(
        license_tag="L-1", name="CANoe Runtime", manufacturer="Vector",
        category="Floating License", quantity=2, purchase_date="2025-07-02",
        termination_date="2028-07-02", purchase_cost=7157.35,
        depreciation="Leasing", maintained=True,
    ))
    assert resp.status_code == 201, resp.text
    leased = resp.json()
    assert leased["quantity"] == 2
    assert leased["maintained"] is True
    # The golden workbook row, rounded the way the grid renders it
    assert leased["per_year"] == {
        "2025": 1192.89, "2026": 2385.78, "2027": 2385.78, "2028": 1391.71,
    }
    assert leased["total"] == pytest.approx(sum(leased["per_year"].values()))
    assert leased["total"] == 7356.16

    resp = client.post(f"/api/hw/projects/{licenses_project}/licenses", json=hw_license(
        license_tag="L-2", name="Planned Compiler", category="Compiler",
        purchase_date="2027-03-01", purchase_cost=999.0,
        depreciation="Planned Purchase",
    ))
    assert resp.status_code == 201
    planned = resp.json()
    assert planned["per_year"] == {"2025": 0.0, "2026": 0.0, "2027": 0.0, "2028": 0.0}
    assert planned["total"] == 0.0

    resp = client.put(f"/api/hw/licenses/{planned['id']}", json=hw_license(
        license_tag="L-2", name="Bought Compiler", category="Compiler",
        purchase_date="2027-03-01", purchase_cost=999.0, depreciation="Purchase",
    ))
    assert resp.status_code == 200
    assert resp.json()["per_year"]["2027"] == 999.0

    assert client.delete(f"/api/hw/licenses/{planned['id']}").status_code == 204
    assert len(client.get(f"/api/hw/projects/{licenses_project}/licenses").json()) == 1

    resp = client.put("/api/hw/licenses/987654", json=hw_license(name="X"))
    assert resp.status_code == 404
    assert resp.json()["detail"] == "Hardware license not found"
    assert client.delete("/api/hw/licenses/987654").status_code == 404


def test_license_bulk_replace(client, licenses_project):
    resp = client.put(f"/api/hw/projects/{licenses_project}/licenses", json={"items": [
        hw_license(license_tag="C-1", name="Compiler Seat", category="Compiler",
                   purchase_date="2026-01-01", termination_date="2028-12-31",
                   purchase_cost=3600.0, depreciation="Leasing"),
    ]})
    assert resp.status_code == 200, resp.text
    rows = resp.json()
    assert len(rows) == 1
    assert rows[0]["per_year"] == {"2026": 1200.0, "2027": 1200.0, "2028": 1200.0}

    resp = client.put(f"/api/hw/projects/{licenses_project}/licenses", json={"items": []})
    assert resp.status_code == 200
    assert client.get(f"/api/hw/projects/{licenses_project}/licenses").json() == []


# ---------------------------------------------------------------------------
# Budget summary
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def summary_project(client):
    project_id = make_project(client, "Summary Project",
                              budget_assets=10000.0, budget_licenses=5000.0)
    resp = client.put(f"/api/hw/projects/{project_id}/assets", json={"items": [
        asset(name="Bench PC", category="PC", status="In Stock",
              purchase_date="2026-02-01", purchase_cost=2000.0,
              purchase_type="Purchase"),
        asset(name="Spare PC", category="PC", status="Labeled",
              purchase_date="2027-04-01", purchase_cost=1500.0,
              purchase_type="Planned Purchase"),
    ]})
    assert resp.status_code == 200, resp.text
    resp = client.put(f"/api/hw/projects/{project_id}/licenses", json={"items": [
        hw_license(name="Debugger Seat", category="Debugger License",
                   purchase_date="2026-06-01", purchase_cost=800.0,
                   depreciation="Purchase"),
        hw_license(name="Future Seat", category="Debugger License",
                   purchase_date="2026-09-01", purchase_cost=300.0,
                   depreciation="Planned Purchase"),
    ]})
    assert resp.status_code == 200, resp.text
    return project_id


def test_adjustments_replace(client, summary_project):
    items = [
        {"year": 2026, "kind": "assets", "amount": 250.0, "note": "Rush order"},
        {"year": 2027, "kind": "licenses", "amount": 125.0, "note": ""},
    ]
    resp = client.put(f"/api/hw/projects/{summary_project}/adjustments",
                      json={"items": items})
    assert resp.status_code == 200, resp.text
    assert resp.json() == items

    # Replacing reuses the (year, kind) pairs: the deletes must land first
    resp = client.put(f"/api/hw/projects/{summary_project}/adjustments",
                      json={"items": items})
    assert resp.status_code == 200, resp.text
    assert client.get(f"/api/hw/projects/{summary_project}/adjustments").json() == items

    resp = client.put(f"/api/hw/projects/{summary_project}/adjustments", json={"items": [
        {"year": 2026, "kind": "assets", "amount": 1.0, "note": ""},
        {"year": 2026, "kind": "assets", "amount": 2.0, "note": ""},
    ]})
    assert resp.status_code == 422
    assert "Duplicate assets adjustment" in resp.text

    resp = client.put(f"/api/hw/projects/{summary_project}/adjustments", json={"items": [
        {"year": 2026, "kind": "hardware", "amount": 1.0, "note": ""},
    ]})
    assert resp.status_code == 422

    # The rejected payloads left the stored adjustments alone
    assert client.get(f"/api/hw/projects/{summary_project}/adjustments").json() == items


def test_project_summary(client, summary_project):
    resp = client.get(f"/api/hw/projects/{summary_project}/summary")
    assert resp.status_code == 200, resp.text
    data = resp.json()

    assert [row["year"] for row in data["years"]] == [2026, 2027]
    by_year = {row["year"]: row for row in data["years"]}

    # 2026: purchased asset 2000 + the 250 special case, purchased license 800,
    # and the planned license 300 kept out of the actual columns
    assert by_year[2026]["actual_assets"] == 2250.0
    assert by_year[2026]["actual_licenses"] == 800.0
    assert by_year[2026]["actual_total"] == 3050.0
    assert by_year[2026]["planned_assets"] == 0.0
    assert by_year[2026]["planned_licenses"] == 300.0
    assert by_year[2026]["planned_total"] == 300.0
    assert by_year[2026]["grand_total"] == 3350.0

    # 2027: only the licenses special case is actual, the planned asset is not
    assert by_year[2027]["actual_assets"] == 0.0
    assert by_year[2027]["actual_licenses"] == 125.0
    assert by_year[2027]["planned_assets"] == 1500.0
    assert by_year[2027]["planned_licenses"] == 0.0
    assert by_year[2027]["grand_total"] == 1625.0

    totals = data["totals"]
    assert totals["actual_assets"] == 2250.0
    assert totals["actual_licenses"] == 925.0
    assert totals["actual_total"] == 3175.0
    assert totals["planned_total"] == 1800.0
    assert totals["grand_total"] == 4975.0

    dashboard = data["dashboard"]
    assert dashboard["budget_total"] == 15000.0
    assert dashboard["spent_total"] == 3175.0
    assert dashboard["planned_total"] == 1800.0
    # The sheet's C7 = C5 - C6: the plan is reported but never subtracted
    assert dashboard["remaining"] == dashboard["budget_total"] - dashboard["spent_total"]
    assert dashboard["remaining"] == 11825.0

    assert data["asset_count"] == 2
    assert data["license_count"] == 2
    assert data["adjustments"] == [
        {"year": 2026, "kind": "assets", "amount": 250.0, "note": "Rush order"},
        {"year": 2027, "kind": "licenses", "amount": 125.0, "note": ""},
    ]

    pivot = data["asset_pivot"]
    assert pivot["statuses"] == ["In Stock", "Labeled"]
    assert pivot["rows"] == [
        {"category": "PC", "counts": {"In Stock": 1, "Labeled": 1}, "total": 2},
    ]

    rollup = next(p for p in client.get("/api/hw/projects").json()
                  if p["id"] == summary_project)
    assert rollup["actual_total"] == 3175.0
    assert rollup["planned_total"] == 1800.0
    assert rollup["remaining"] == 11825.0


def test_summary_adjustment_only_year(client):
    """A special case booked outside the registers still gets its own row."""
    project_id = make_project(client, "Adjustment Only", budget_assets=500.0)
    resp = client.put(f"/api/hw/projects/{project_id}/adjustments", json={"items": [
        {"year": 2029, "kind": "assets", "amount": 400.0, "note": "Carry over"},
    ]})
    assert resp.status_code == 200, resp.text

    data = client.get(f"/api/hw/projects/{project_id}/summary").json()
    assert [row["year"] for row in data["years"]] == [2029]
    assert data["years"][0]["actual_assets"] == 400.0
    assert data["dashboard"]["remaining"] == 100.0

    assert client.delete(f"/api/hw/projects/{project_id}").status_code == 204


# ---------------------------------------------------------------------------
# Renewal risk
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def risk_project(client):
    today = date.today()

    def in_days(days):
        return (today + timedelta(days=days)).isoformat()

    project_id = make_project(client, "Renewal Risk")
    resp = client.put(f"/api/hw/projects/{project_id}/licenses", json={"items": [
        hw_license(name="Lapsed Seat", manufacturer="Vector",
                   expiration_date=in_days(-10)),
        hw_license(name="Due Seat", manufacturer="Lauterbach",
                   expiration_date=in_days(5)),
        hw_license(name="Next Quarter Seat", manufacturer="ETAS",
                   expiration_date=in_days(45)),
        hw_license(name="Late Quarter Seat", manufacturer="dSPACE",
                   expiration_date=in_days(75)),
        hw_license(name="Far Away Seat", manufacturer="Vector",
                   expiration_date=in_days(200)),
        hw_license(name="Perpetual Seat", manufacturer="MathWorks"),
    ]})
    assert resp.status_code == 200, resp.text
    return project_id


def test_renewal_risk_buckets(client, risk_project):
    data = client.get(f"/api/hw/projects/{risk_project}/summary").json()
    assert data["risk"] == {
        "expired": 1, "in_30_days": 1, "in_60_days": 1, "in_90_days": 1,
    }

    # Everything inside the 90-day horizon, soonest first; the +200 day and the
    # perpetual seat are out
    expiring = data["expiring"]
    assert [row["name"] for row in expiring] == [
        "Lapsed Seat", "Due Seat", "Next Quarter Seat", "Late Quarter Seat",
    ]
    assert [row["days_left"] for row in expiring] == [-10, 5, 45, 75]
    assert expiring[0]["manufacturer"] == "Vector"
    assert all(row["hw_project_id"] == risk_project for row in expiring)
    assert all(row["hw_project_name"] == "Renewal Risk" for row in expiring)

    rollup = next(p for p in client.get("/api/hw/projects").json()
                  if p["id"] == risk_project)
    assert rollup["licenses_expired"] == 1
    assert rollup["licenses_expiring_90"] == 3

    # Expiry dates alone never widen the budget span
    assert [row["year"] for row in data["years"]] == [date.today().year]


# ---------------------------------------------------------------------------
# Workbook import / export
# ---------------------------------------------------------------------------

IMPORT_ASSET_ROWS = [
    ["A-1", "Vehiclevo", "Lauterbach Trace32", "SN-1", "PowerDebug",
     "Lauterbach debugger", "In Stock", "Lauterbach", "2026-03-15", "1.234,56",
     "PO-9", "15.03.2029", "M. Hegazy", "", "Procurement", "Purchase", 0, "junk"],
    # No name but real data: skipped with a warning naming the row
    ["", "Ghost", "", "", "", "", "", "", "", 500, "", "", "", "", "",
     "Not Purchased", 0, ""],
]

IMPORT_LICENSE_ROWS = [
    ["L-1", "Vehiclevo", "CANoe Runtime", "KEY-123", "2027-01-31",
     "fleet@vehiclevo.com", "Floating License", "Vector", "Vector", 2,
     "2026-01-01", "2028-12-31", "Leasing", "yes", "€ 3600", "PO-1", "Imported"],
    ["L-2", "Vehiclevo", "Legacy Tool", "", "", "", "Maintenance", "Vector",
     "Vector", 1, "", "", "Rental", "no", 0, "", ""],
]


@pytest.fixture(scope="module")
def import_project(client):
    return make_project(client, "Import Project", budget_assets=5000.0,
                        budget_licenses=5000.0)


@pytest.fixture(scope="module")
def import_workbook():
    return build_workbook([
        # Trailing year and junk columns exercise the "derived" and "unknown"
        # header rules
        ("Assets", hw_excel.ASSET_HEADERS + ["2026", "Rubbish"], IMPORT_ASSET_ROWS),
        ("Licenses", hw_excel.LICENSE_HEADERS, IMPORT_LICENSE_ROWS),
    ])


def upload(client, project_id, content, dry_run, name="register.xlsx"):
    return client.post(
        f"/api/hw/projects/{project_id}/import",
        params={"dry_run": str(dry_run).lower()},
        files={"file": (name, content, XLSX_MEDIA_TYPE)},
    )


# ---------------------------------------------------------------------------
# Budget mode
# ---------------------------------------------------------------------------

def test_budget_defaults_to_the_split_of_assets_and_licenses(client):
    project_id = make_project(client, "Split Budget",
                              budget_assets=1000.0, budget_licenses=250.0)
    project = client.get(f"/api/hw/projects/{project_id}").json()
    assert project["budget_mode"] == "split"

    dashboard = client.get(f"/api/hw/projects/{project_id}/summary").json()["dashboard"]
    assert dashboard["budget_total"] == 1250.0
    assert dashboard["budget_assets"] == 1000.0
    assert dashboard["budget_licenses"] == 250.0


def test_overall_budget_ignores_the_split_figures(client):
    """One approved number is the whole budget; a stale split must not add to it."""
    project_id = make_project(
        client, "Overall Budget", budget_mode="overall", budget_total=9000.0,
        # Left over from a previous split — these must not count
        budget_assets=1000.0, budget_licenses=250.0,
    )
    dashboard = client.get(f"/api/hw/projects/{project_id}/summary").json()["dashboard"]
    assert dashboard["budget_total"] == 9000.0
    assert dashboard["budget_assets"] == 0.0
    assert dashboard["budget_licenses"] == 0.0
    assert dashboard["remaining"] == 9000.0

    rollup = next(p for p in client.get("/api/hw/projects").json()
                  if p["id"] == project_id)
    assert rollup["budget_total"] == 9000.0


def test_switching_budget_mode_reinstates_the_split(client):
    project_id = make_project(client, "Mode Switch", budget_mode="overall",
                              budget_total=9000.0, budget_assets=600.0,
                              budget_licenses=400.0)
    body = client.get(f"/api/hw/projects/{project_id}").json()
    body["budget_mode"] = "split"
    assert client.put(f"/api/hw/projects/{project_id}", json=body).status_code == 200

    dashboard = client.get(f"/api/hw/projects/{project_id}/summary").json()["dashboard"]
    assert dashboard["budget_total"] == 1000.0


def test_unknown_budget_mode_is_rejected(client):
    resp = client.post("/api/hw/projects", json={
        "name": "Bad Mode", "company": "", "description": "",
        "budget_mode": "guess", "budget_total": 0.0,
        "budget_assets": 0.0, "budget_licenses": 0.0, "portal_reference": "",
    })
    assert resp.status_code == 422


def test_overview_totals_overall_and_split_projects_together(client):
    # The suite shares one database, so assert on the delta these two add.
    before = client.get("/api/hw/overview").json()["dashboard"]

    split_id = make_project(client, "Overview Split", budget_assets=100.0,
                            budget_licenses=200.0)
    overall_id = make_project(client, "Overview Overall", budget_mode="overall",
                              budget_total=700.0)

    overview = client.get("/api/hw/overview").json()
    assert {split_id, overall_id} <= {p["id"] for p in overview["projects"]}
    after = overview["dashboard"]

    assert after["budget_total"] - before["budget_total"] == 1000.0
    # The overall project adds to the total but has no per-type share to add, so
    # only the split project moves the breakdown.
    assert after["budget_assets"] - before["budget_assets"] == 100.0
    assert after["budget_licenses"] - before["budget_licenses"] == 200.0

    for pid in (split_id, overall_id):
        assert client.delete(f"/api/hw/projects/{pid}").status_code == 204


def test_import_names_rows_the_working_document_left_blank(client, import_project):
    """The real document names most licences only by category + manufacturer.

    Those rows carry dates and money, so skipping them for want of a name would
    throw the register away; they are named from what is there instead.
    """
    workbook = build_workbook([
        ("Licenses", hw_excel.LICENSE_HEADERS, [
            # Exactly the shape of the customer's sheet: no Name, no Company
            ["", "", "", "", "2026-07-09", "", "Compiler", "", "Tasking", 2,
             "2025-07-02", "2026-07-09", "Purchase", "", 9877, "", ""],
            # Nothing identifying at all — still skipped
            ["", "Ghost Corp", "", "", "", "", "", "", "", 1, "", "", "", "", 0, "", ""],
        ]),
    ])
    data = upload(client, import_project, workbook, dry_run=True).json()

    assert len(data["licenses"]) == 1
    named = data["licenses"][0]
    assert named["name"] == "Compiler — Tasking"
    assert named["purchase_cost"] == 9877.0
    assert any("named 'Compiler — Tasking'" in w for w in data["warnings"])
    assert any("skipped, no Name" in w for w in data["warnings"])


def test_import_dry_run(client, import_project, import_workbook):
    resp = upload(client, import_project, import_workbook, dry_run=True)
    assert resp.status_code == 200, resp.text
    data = resp.json()

    assert data["sheets_found"] == ["Assets", "Licenses"]
    assert len(data["assets"]) == 1
    assert len(data["licenses"]) == 2

    imported = data["assets"][0]
    assert imported["name"] == "Lauterbach Trace32"
    assert imported["purchase_date"] == "2026-03-15"
    assert imported["eol_date"] == "2029-03-15"
    assert imported["purchase_cost"] == 1234.56
    assert imported["purchase_type"] == "Purchase"
    assert imported["catalog_item_id"] is None

    seat = data["licenses"][0]
    assert seat["quantity"] == 2
    assert seat["maintained"] is True
    assert seat["purchase_cost"] == 3600.0
    assert seat["depreciation"] == "Leasing"

    warnings = data["warnings"]
    assert any("skipped, no Asset Name" in w for w in warnings)
    assert any("ignored unknown column 'Rubbish'" in w for w in warnings)
    assert any("unknown Depreciation 'Rental'" in w for w in warnings)
    # A 4-digit header is a derived year column, not an unknown one
    assert not any("'2026'" in w for w in warnings)

    # Nothing was written
    assert client.get(f"/api/hw/projects/{import_project}/assets").json() == []


def test_import_writes_rows(client, import_project, import_workbook):
    resp = upload(client, import_project, import_workbook, dry_run=False)
    assert resp.status_code == 200, resp.text
    assert resp.json()["created_assets"] == 1
    assert resp.json()["created_licenses"] == 2

    assets = client.get(f"/api/hw/projects/{import_project}/assets").json()
    assert len(assets) == 1
    assert assets[0]["asset_tag"] == "A-1"
    assert assets[0]["purchase_cost"] == 1234.56
    # Span runs from the earliest purchase to the latest EOL across both sheets
    assert assets[0]["per_year"] == {
        "2026": 1234.56, "2027": 0.0, "2028": 0.0, "2029": 0.0,
    }

    licenses = client.get(f"/api/hw/projects/{import_project}/licenses").json()
    assert [row["license_tag"] for row in licenses] == ["L-1", "L-2"]
    assert licenses[0]["per_year"] == {
        "2026": 1200.0, "2027": 1200.0, "2028": 1200.0, "2029": 0.0,
    }
    assert licenses[0]["total"] == 3600.0
    # The unknown depreciation fell back to the inert default
    assert licenses[1]["depreciation"] == "Not Purchased"
    assert licenses[1]["total"] == 0.0


def test_import_rejects_non_workbook(client, import_project):
    resp = upload(client, import_project, b"not a workbook at all", dry_run=True,
                  name="notes.txt")
    assert resp.status_code == 400
    assert "could not read workbook" in resp.json()["detail"]

    resp = upload(client, import_project,
                  build_workbook([("Contacts", ["Name"], [["Nobody"]])]),
                  dry_run=True)
    assert resp.status_code == 400
    assert resp.json()["detail"] == "no Assets or Licenses sheet found"

    resp = upload(client, 987654, b"x", dry_run=True)
    assert resp.status_code == 404


def test_import_template_download(client):
    resp = client.get("/api/hw/import-template.xlsx")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(XLSX_MEDIA_TYPE)
    assert "attachment;" in resp.headers["content-disposition"]

    workbook = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert {"Assets", "Licenses"} <= set(workbook.sheetnames)
    header = [cell.value for cell in workbook["Assets"][1]]
    assert header == hw_excel.ASSET_HEADERS
    header = [cell.value for cell in workbook["Licenses"][1]]
    assert header == hw_excel.LICENSE_HEADERS


def test_import_template_round_trips(client, import_project):
    """The blank template parses to an empty, warning-free register."""
    template = client.get("/api/hw/import-template.xlsx").content
    resp = upload(client, import_project, template, dry_run=True,
                  name="template.xlsx")
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["assets"] == []
    assert data["licenses"] == []
    assert data["warnings"] == []


def test_project_export_xlsx(client, import_project):
    resp = client.get(f"/api/hw/projects/{import_project}/export.xlsx")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(XLSX_MEDIA_TYPE)
    assert "Import Project" in resp.headers["content-disposition"]
    assert resp.content[:2] == b"PK"

    workbook = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert workbook.sheetnames == [
        "Dashboard", "Summary", "Assets", "Licenses", "HW Catalogue",
    ]
    sheet = workbook["Assets"]
    header = [cell.value for cell in sheet[1]]
    assert header[:len(hw_excel.ASSET_HEADERS)] == hw_excel.ASSET_HEADERS
    # The register keeps the derived year columns the sheet showed
    assert header[len(hw_excel.ASSET_HEADERS):] == ["2026", "2027", "2028", "2029"]

    names = [row[2].value for row in sheet.iter_rows(min_row=2)]
    assert "Lauterbach Trace32" in names
    # ... and closes with the footer the sheet had, summing the year columns
    footer = [cell.value for cell in sheet[sheet.max_row]]
    assert footer[0] == "TOTAL"
    assert footer[len(hw_excel.ASSET_HEADERS):] == [1234.56, 0.0, 0.0, 0.0]


def test_empty_project_summary_and_export(client):
    """A brand-new project still summarises and exports, on the current year."""
    project_id = make_project(client, "Fresh Project", budget_assets=1000.0)

    data = client.get(f"/api/hw/projects/{project_id}/summary").json()
    assert [row["year"] for row in data["years"]] == [date.today().year]
    assert data["totals"]["grand_total"] == 0.0
    assert data["expiring"] == []
    assert data["asset_pivot"] == {"statuses": [], "rows": []}
    assert data["dashboard"]["remaining"] == 1000.0

    resp = client.get(f"/api/hw/projects/{project_id}/export.xlsx")
    assert resp.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert workbook["Assets"]["A2"].value == "TOTAL"

    assert client.delete(f"/api/hw/projects/{project_id}").status_code == 204


def test_export_round_trips_back_into_a_register(client, import_workbook):
    """Exporting and re-importing reproduces the rows, footer row and all."""
    project_id = make_project(client, "Round Trip")
    assert upload(client, project_id, import_workbook, dry_run=False).status_code == 200

    exported = client.get(f"/api/hw/projects/{project_id}/export.xlsx").content
    target = make_project(client, "Round Trip Target")
    resp = upload(client, target, exported, dry_run=True, name="export.xlsx")
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert len(data["assets"]) == 1
    assert len(data["licenses"]) == 2
    assert data["warnings"] == []
    assert data["assets"][0]["purchase_cost"] == 1234.56

    for stale in (project_id, target):
        assert client.delete(f"/api/hw/projects/{stale}").status_code == 204


# ---------------------------------------------------------------------------
# Management overview
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def overview_projects(client):
    """Two projects parked in years no other test uses, so their rows stand alone."""
    first = make_project(client, "Overview A", budget_assets=1000.0)
    resp = client.put(f"/api/hw/projects/{first}/assets", json={"items": [
        asset(name="Overview Bench", category="PC", status="In Stock",
              purchase_date="2031-06-01", purchase_cost=700.0,
              purchase_type="Purchase"),
    ]})
    assert resp.status_code == 200, resp.text

    second = make_project(client, "Overview B", budget_licenses=2000.0)
    resp = client.put(f"/api/hw/projects/{second}/licenses", json={"items": [
        hw_license(name="Overview Seat", category="Compiler",
                   purchase_date="2032-06-01", purchase_cost=900.0,
                   depreciation="Purchase"),
        hw_license(name="Overview Plan", category="Compiler",
                   purchase_date="2031-02-01", purchase_cost=400.0,
                   depreciation="Planned Purchase"),
    ]})
    assert resp.status_code == 200, resp.text
    return first, second


def test_overview_aggregates_projects(client, overview_projects):
    first, second = overview_projects
    resp = client.get("/api/hw/overview")
    assert resp.status_code == 200, resp.text
    data = resp.json()

    rollups = client.get("/api/hw/projects").json()
    assert data["project_count"] == len(rollups)
    assert [p["id"] for p in data["projects"]] == [p["id"] for p in rollups]
    assert data["asset_count"] == sum(p["asset_count"] for p in rollups)
    assert data["license_count"] == sum(p["license_count"] for p in rollups)
    assert data["dashboard"]["budget_total"] == pytest.approx(
        sum(p["budget_total"] for p in rollups))

    by_id = {p["id"]: p for p in data["projects"]}
    assert by_id[first]["asset_count"] == 1
    assert by_id[first]["actual_total"] == 700.0
    assert by_id[first]["remaining"] == 300.0
    assert by_id[second]["license_count"] == 2
    assert by_id[second]["actual_total"] == 900.0
    assert by_id[second]["planned_total"] == 400.0

    # 2031/2032 belong to these two projects alone, so the merged rows are exact
    by_year = {row["year"]: row for row in data["years"]}
    assert by_year[2031]["actual_assets"] == 700.0
    assert by_year[2031]["actual_licenses"] == 0.0
    assert by_year[2031]["planned_licenses"] == 400.0
    assert by_year[2031]["grand_total"] == 1100.0
    assert by_year[2032]["actual_licenses"] == 900.0
    assert by_year[2032]["actual_total"] == 900.0

    # The span is contiguous across every project's rows
    years = [row["year"] for row in data["years"]]
    assert years == list(range(years[0], years[-1] + 1))
    assert data["totals"]["actual_total"] == pytest.approx(
        data["dashboard"]["spent_total"])

    # Expiring licenses are labelled with the project they belong to
    assert any(row["hw_project_name"] == "Renewal Risk" for row in data["expiring"])
    assert data["risk"]["expired"] >= 1


@pytest.fixture()
def cascade_project(client):
    project_id = make_project(client, "Cascade Delete", budget_assets=100.0)
    resp = client.post(f"/api/hw/projects/{project_id}/assets",
                       json=asset(name="Doomed Asset", purchase_cost=10.0))
    assert resp.status_code == 201, resp.text
    resp = client.post(f"/api/hw/projects/{project_id}/licenses",
                       json=hw_license(name="Doomed Seat", purchase_cost=10.0))
    assert resp.status_code == 201, resp.text
    resp = client.put(f"/api/hw/projects/{project_id}/adjustments", json={"items": [
        {"year": 2030, "kind": "assets", "amount": 5.0, "note": ""},
    ]})
    assert resp.status_code == 200, resp.text
    return project_id


def test_overview_after_project_delete(client, cascade_project):
    """Deleting a project takes its registers and its rollup with it."""
    project_id = cascade_project
    before = client.get("/api/hw/overview").json()
    assert any(p["id"] == project_id for p in before["projects"])

    assert client.delete(f"/api/hw/projects/{project_id}").status_code == 204
    after = client.get("/api/hw/overview").json()
    assert not any(p["id"] == project_id for p in after["projects"])
    assert after["project_count"] == before["project_count"] - 1
    assert after["asset_count"] == before["asset_count"] - 1
    assert after["license_count"] == before["license_count"] - 1
