"""Regressions for the correctness fixes of the September 2026 review (Phase 1).

Each test names the review finding it guards. They run against the shared
SQLite test database like the rest of the suite and create every row they use.
"""

import io
import logging
from datetime import date, timedelta

import openpyxl
import pytest
from sqlalchemy import create_engine, text

from app import models
from app.config import DATE_WINDOW_YEARS
from app.database import SessionLocal, enable_sqlite_foreign_keys, run_migrations
from app.routers.hardware import XLSX_MEDIA_TYPE
from app.services import hw_depreciation as dep
from app.services import hw_excel
from app.services.http import attachment_disposition
from test_hw_management import asset, build_workbook, hw_license, make_project, upload


def new_project(client, name="Fixes", start=(2026, 1), end=(2027, 12), **extra):
    resp = client.post("/api/projects", json={
        "name": name, "company": "Vehiclevo",
        "start_year": start[0], "start_month": start[1],
        "end_year": end[0], "end_month": end[1],
        **extra,
    })
    assert resp.status_code == 201, resp.text
    return resp.json()


def new_feature(client, project_id, name="Feature"):
    resp = client.post(f"/api/projects/{project_id}/features", json={"name": name})
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


def new_role(client, feature_id, **overrides):
    body = {
        "name": "Developer", "location": "BCC", "level": "Senior", "ftes": 1.0,
        "use_advanced_allocation": False, "allocations": [],
        **overrides,
    }
    resp = client.post(f"/api/features/{feature_id}/roles", json=body)
    assert resp.status_code == 201, resp.text
    return resp.json()


def updated_at(client, project_id):
    return client.get(f"/api/projects/{project_id}").json()["updated_at"]


# ---------------------------------------------------------------------------
# F-01: download filenames
# ---------------------------------------------------------------------------

def test_attachment_disposition_is_latin1_safe_and_carries_the_real_name():
    header = attachment_disposition("Straße – Plan ✓.xlsx")
    header.encode("latin-1")  # the header must be encodable, whatever the name
    assert header.startswith('attachment; filename="')
    assert "filename*=UTF-8''Stra" in header
    assert "%E2%80%93" in header  # the en dash travels percent-encoded
    # Quotes and path separators never reach the ASCII fallback
    fallback = attachment_disposition('a"b/c\\d.xlsx').split(";")[1]
    assert '"' not in fallback.split("=", 1)[1].strip('" ')
    assert "/" not in fallback and "\\" not in fallback
    # A name with nothing ASCII in it still gets a usable fallback
    assert 'filename="download"' in attachment_disposition("✓✓✓")


def test_exports_with_non_latin_names_download(client):
    project = new_project(client, name="Übersicht – Ärger ✓")
    feature_id = new_feature(client, project["id"])
    new_role(client, feature_id)
    resp = client.get(f"/api/projects/{project['id']}/reports/resource-plan.xlsx")
    assert resp.status_code == 200, resp.text
    assert "filename*=UTF-8''" in resp.headers["content-disposition"]

    hw_id = make_project(client, "Größe ✓")
    resp = client.get(f"/api/hw/projects/{hw_id}/export.xlsx")
    assert resp.status_code == 200, resp.text
    assert "filename*=UTF-8''Gr%C3%B6%C3%9Fe" in resp.headers["content-disposition"]


# ---------------------------------------------------------------------------
# F-04: shrinking the timeline
# ---------------------------------------------------------------------------

def test_timeline_change_refuses_to_orphan_periods_and_hardware(client):
    project = new_project(client, start=(2026, 1), end=(2027, 12))
    pid = project["id"]
    feature_id = new_feature(client, pid, "ADAS")
    new_role(client, feature_id, name="Architect", ftes=0.0, use_advanced_allocation=True,
             allocations=[{"start_month": "2027-01", "end_month": "2027-12", "ftes": 1.0}])
    resp = client.post(f"/api/projects/{pid}/hardware", json={
        "name": "Debugger", "billing": "once", "unit_cost": 100.0, "qty": 1,
        "years": [2027],
    })
    assert resp.status_code == 201, resp.text

    resp = client.put(f"/api/projects/{pid}", json={"end_year": 2026, "end_month": 12})
    assert resp.status_code == 422
    detail = resp.json()["detail"]
    assert "ADAS / Architect: period 2027-01 to 2027-12" in detail
    assert "hardware item Debugger: 2027" in detail
    # Nothing moved
    assert client.get(f"/api/projects/{pid}").json()["end_year"] == 2027

    # Widening is always fine, and a change that keeps the rows inside is too
    resp = client.put(f"/api/projects/{pid}", json={"end_year": 2028, "end_month": 6})
    assert resp.status_code == 200, resp.text
    resp = client.put(f"/api/projects/{pid}", json={"start_year": 2026, "start_month": 6})
    assert resp.status_code == 200, resp.text


def test_timeline_change_drops_quotas_for_years_no_longer_covered(client):
    project = new_project(client, start=(2026, 1), end=(2027, 12))
    pid = project["id"]
    resp = client.put(f"/api/projects/{pid}/rates", json={
        "ticket_quotas": {"2026": {"small": 10, "medium": 10, "large": 10},
                          "2027": {"small": 20, "medium": 20, "large": 20}},
    })
    assert resp.status_code == 200, resp.text
    resp = client.put(f"/api/projects/{pid}", json={"end_year": 2026, "end_month": 12})
    assert resp.status_code == 200, resp.text
    quotas = client.get(f"/api/projects/{pid}/rates").json()["ticket_quotas"]
    assert list(quotas) == ["2026"]
    new_role(client, new_feature(client, pid))
    assert client.get(f"/api/projects/{pid}/validate").json()["valid"] is True


def test_validation_reports_rows_outside_the_timeline():
    """Rows written before the guard existed are reported, not hidden."""
    db = SessionLocal()
    try:
        project = models.Project(name="Legacy", company="Co", start_year=2026,
                                 start_month=1, end_year=2026, end_month=12)
        db.add(project)
        db.flush()
        db.add(models.HardwareItem(project_id=project.id, name="Old rig", years_json="[2030]"))
        db.commit()
        db.refresh(project)
        from app.services.calculations import validate_project
        errors = validate_project(project)
        assert any("Outside the project timeline: hardware item Old rig: 2030" in e
                   for e in errors)
        db.delete(project)
        db.commit()
    finally:
        db.close()


# ---------------------------------------------------------------------------
# F-05: updated_at follows child rows
# ---------------------------------------------------------------------------

def test_project_updated_at_moves_with_its_children(client):
    project = new_project(client)
    pid = project["id"]
    stamp = updated_at(client, pid)

    feature_id = new_feature(client, pid)
    after_feature = updated_at(client, pid)
    assert after_feature > stamp

    role = new_role(client, feature_id)
    after_role = updated_at(client, pid)
    assert after_role >= after_feature

    resp = client.put(f"/api/roles/{role['id']}", json={
        "name": "Developer", "location": "BCC", "level": "Senior", "ftes": 0.0,
        "use_advanced_allocation": True,
        "allocations": [{"start_month": "2026-01", "end_month": "2026-06", "ftes": 1.5}],
    })
    assert resp.status_code == 200, resp.text
    after_allocation = updated_at(client, pid)
    assert after_allocation >= after_role

    resp = client.put(f"/api/projects/{pid}/rates", json={"sp_to_hours": 5.0})
    assert resp.status_code == 200, resp.text
    after_rates = updated_at(client, pid)
    assert after_rates >= after_allocation

    resp = client.put(f"/api/projects/{pid}/financial-data",
                      json={"encrypted_money": "AAAA", "money_iv": "BBBB"})
    assert resp.status_code == 200, resp.text
    after_blob = updated_at(client, pid)
    assert after_blob >= after_rates

    resp = client.delete(f"/api/features/{feature_id}")
    assert resp.status_code == 204
    assert updated_at(client, pid) >= after_blob


def test_hw_project_updated_at_moves_with_its_registers(client):
    hw_id = make_project(client, "Touched")
    before = client.get(f"/api/hw/projects/{hw_id}").json()["updated_at"]
    resp = client.post(f"/api/hw/projects/{hw_id}/assets", json=asset(name="Probe"))
    assert resp.status_code == 201, resp.text
    assert client.get(f"/api/hw/projects/{hw_id}").json()["updated_at"] > before


# ---------------------------------------------------------------------------
# F-06 / F-38: the date window
# ---------------------------------------------------------------------------

def test_date_window_keeps_typos_out_of_the_year_span():
    typo = hw_license(name="Typo", purchase_date=date(225, 7, 2), termination_date=None,
                      purchase_cost=1.0, depreciation="Purchase")
    real = hw_license(name="Real", purchase_date=date(2025, 7, 2),
                      termination_date=date(2028, 7, 2), purchase_cost=1.0,
                      depreciation="Leasing")

    class Row:
        def __init__(self, **kw):
            self.__dict__.update(kw)

    years = dep.year_span([], [Row(**typo), Row(**real)])
    assert years == [2025, 2026, 2027, 2028]
    assert dep.year_span([], [], extra_years=[1, 2026, 99999]) == [2026]
    first, last = DATE_WINDOW_YEARS
    assert (first, last) == (1990, 2100)


def test_importer_refuses_serials_and_years_outside_the_window():
    workbook = build_workbook([
        ("Assets", hw_excel.ASSET_HEADERS, [
            # A bare year typed into the purchase date column is not a 1905 serial
            ["A-1", "Co", "Bare year", "", "", "", "", "", 2026, 100, "", "", "", "", "",
             "Purchase"],
            # A real serial (2026-03-15) still reads as a date
            ["A-2", "Co", "Serial", "", "", "", "", "", 46096, 100, "", "", "", "", "",
             "Purchase"],
            # A year 0225 typo is reported and left empty
            ["A-3", "Co", "Typo", "", "", "", "", "", "0225-07-02", 100, "", "", "", "", "",
             "Purchase"],
        ]),
    ])
    parsed = hw_excel.parse_workbook(workbook)
    by_tag = {row["asset_tag"]: row for row in parsed["assets"]}
    assert by_tag["A-1"]["purchase_date"] is None
    assert by_tag["A-2"]["purchase_date"] == "2026-03-15"
    assert by_tag["A-3"]["purchase_date"] is None
    assert sum("is not a date between 1990 and 2100" in w for w in parsed["warnings"]) == 2


# ---------------------------------------------------------------------------
# F-07: SQLite enforces foreign keys
# ---------------------------------------------------------------------------

def test_sqlite_cascades_and_set_null_apply(client):
    base = new_project(client, name="Base")
    resp = client.post(f"/api/projects/{base['id']}/clone",
                       json={"name": "Scenario A", "as_scenario": True})
    assert resp.status_code == 201, resp.text
    scenario_id = resp.json()["id"]

    resp = client.post("/api/hardware-catalog", json={"name": "Vendor probe", "unit_cost": 5.0})
    assert resp.status_code == 201, resp.text
    catalog_id = resp.json()["id"]
    resp = client.post(f"/api/projects/{base['id']}/hardware", json={
        "name": "Vendor probe", "catalog_item_id": catalog_id, "years": [2026],
    })
    assert resp.status_code == 201, resp.text
    item_id = resp.json()["id"]

    assert client.delete(f"/api/hardware-catalog/{catalog_id}").status_code == 204
    db = SessionLocal()
    try:
        assert db.get(models.HardwareItem, item_id).catalog_item_id is None
    finally:
        db.close()

    assert client.delete(f"/api/projects/{base['id']}").status_code == 204
    assert client.get(f"/api/projects/{scenario_id}").status_code == 404
    db = SessionLocal()
    try:
        assert db.get(models.HardwareItem, item_id) is None
    finally:
        db.close()


def test_migrations_run_with_foreign_keys_off_and_leave_them_on(tmp_path):
    fresh = create_engine(f"sqlite:///{tmp_path / 'fk.db'}")
    enable_sqlite_foreign_keys(fresh)
    try:
        run_migrations(fresh)
        with fresh.connect() as connection:
            assert connection.execute(text("PRAGMA foreign_keys")).scalar() == 1
    finally:
        fresh.dispose()


# ---------------------------------------------------------------------------
# F-08: import modes and duplicate warnings
# ---------------------------------------------------------------------------

REGISTER_ROWS = [
    ["A-1", "Co", "Trace32", "SN1", "", "Debugger", "In Use", "Lauterbach",
     "2026-03-15", 1000, "", "2029-03-15", "", "", "", "Purchase"],
]


def test_import_append_and_replace_modes(client):
    hw_id = make_project(client, "Modes")
    workbook = build_workbook([("Assets", hw_excel.ASSET_HEADERS, REGISTER_ROWS)])

    assert upload(client, hw_id, workbook, dry_run=False).status_code == 200
    preview = upload(client, hw_id, workbook, dry_run=True).json()
    assert any("already in the register (A-1)" in w for w in preview["warnings"])

    resp = client.post(f"/api/hw/projects/{hw_id}/import",
                       params={"dry_run": "false", "mode": "append"},
                       files={"file": ("r.xlsx", workbook, XLSX_MEDIA_TYPE)})
    assert resp.status_code == 200, resp.text
    assert resp.json()["replaced_assets"] == 0
    assert len(client.get(f"/api/hw/projects/{hw_id}/assets").json()) == 2

    # Replacing clears the assets register but leaves licenses alone
    resp = client.post(f"/api/hw/projects/{hw_id}/licenses",
                       json=hw_license(name="Kept licence"))
    assert resp.status_code == 201
    resp = client.post(f"/api/hw/projects/{hw_id}/import",
                       params={"dry_run": "false", "mode": "replace"},
                       files={"file": ("r.xlsx", workbook, XLSX_MEDIA_TYPE)})
    assert resp.status_code == 200, resp.text
    assert resp.json()["replaced_assets"] == 2
    assert resp.json()["replaced_licenses"] == 0
    assert len(client.get(f"/api/hw/projects/{hw_id}/assets").json()) == 1
    assert len(client.get(f"/api/hw/projects/{hw_id}/licenses").json()) == 1

    resp = client.post(f"/api/hw/projects/{hw_id}/import",
                       params={"dry_run": "false", "mode": "sideways"},
                       files={"file": ("r.xlsx", workbook, XLSX_MEDIA_TYPE)})
    assert resp.status_code == 422


def test_import_flags_ids_repeated_inside_the_file(client):
    hw_id = make_project(client, "Repeats")
    workbook = build_workbook([("Assets", hw_excel.ASSET_HEADERS,
                                REGISTER_ROWS + REGISTER_ROWS)])
    preview = upload(client, hw_id, workbook, dry_run=True).json()
    assert any("appear more than once in the file (A-1)" in w for w in preview["warnings"])


def test_import_of_a_workbook_without_rows_is_rejected(client):
    hw_id = make_project(client, "Nothing to import")
    workbook = build_workbook([("Assets", hw_excel.ASSET_HEADERS, [])])
    resp = upload(client, hw_id, workbook, dry_run=False)
    assert resp.status_code == 400
    assert "no rows" in resp.json()["detail"]


# ---------------------------------------------------------------------------
# F-37: startup migrations keep logging alive
# ---------------------------------------------------------------------------

def test_startup_migrations_do_not_disable_application_logging(tmp_path):
    logger = logging.getLogger("uvicorn.error")
    logger.disabled = False
    root_level = logging.getLogger().level
    fresh = create_engine(f"sqlite:///{tmp_path / 'log.db'}")
    try:
        run_migrations(fresh)
    finally:
        fresh.dispose()
    assert logger.disabled is False
    assert logging.getLogger().level == root_level


# ---------------------------------------------------------------------------
# F-39: no formula injection
# ---------------------------------------------------------------------------

def test_exports_write_formula_looking_text_as_text(client):
    project = new_project(client, name="=HYPERLINK(\"http://x\")")
    feature_id = new_feature(client, project["id"], "=1+1")
    new_role(client, feature_id, name="=cmd|' /C calc'!A0")
    resp = client.get(f"/api/projects/{project['id']}/reports/resource-plan.xlsx")
    assert resp.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(resp.content))
    formulas = [cell for sheet in workbook for row in sheet.iter_rows() for cell in row
                if cell.data_type == "f"]
    assert formulas == []

    resp = client.post(f"/api/projects/{project['id']}/hardware",
                       json={"name": "=1+1", "supplier_name": "=2+2", "years": [2026]})
    assert resp.status_code == 201, resp.text
    resp = client.get(f"/api/projects/{project['id']}/reports/hardware-plan.xlsx")
    assert resp.status_code == 200
    sheet = openpyxl.load_workbook(io.BytesIO(resp.content))["Hardware Plan"]
    assert sheet["B2"].value == "=1+1" and sheet["B2"].data_type == "s"

    hw_id = make_project(client, "=SUM(A1)")
    resp = client.post(f"/api/hw/projects/{hw_id}/assets", json=asset(name="=3+3"))
    assert resp.status_code == 201
    resp = client.get(f"/api/hw/projects/{hw_id}/export.xlsx")
    workbook = openpyxl.load_workbook(io.BytesIO(resp.content))
    formulas = [cell for sheet in workbook for row in sheet.iter_rows() for cell in row
                if cell.data_type == "f"]
    assert formulas == []


# ---------------------------------------------------------------------------
# F-40: non-finite numbers and bounds
# ---------------------------------------------------------------------------

def test_non_finite_and_absurd_numbers_are_rejected_with_a_readable_422(client):
    project = new_project(client)
    pid = project["id"]
    for body in ('{"sp_to_hours": NaN}', '{"sp_to_hours": Infinity}', '{"sp_to_hours": 1e400}'):
        resp = client.put(f"/api/projects/{pid}/rates", content=body,
                          headers={"content-type": "application/json"})
        assert resp.status_code == 422, body
        detail = resp.json()["detail"]
        assert all(set(entry) == {"loc", "msg", "type"} for entry in detail)

    resp = client.post(f"/api/projects/{pid}/hardware",
                       json={"name": "Big", "qty": 10_000_000, "years": [2026]})
    assert resp.status_code == 422
    resp = client.post(f"/api/projects/{pid}/hardware",
                       json={"name": "Rich", "unit_cost": 1e15, "years": [2026]})
    assert resp.status_code == 422
    resp = client.put(f"/api/projects/{pid}/rates", json={"risk_factor_pct": 2_000_000})
    assert resp.status_code == 422
    resp = client.put(f"/api/projects/{pid}/rates", json={"sp_to_hours": -1})
    assert resp.status_code == 422


def test_validation_errors_do_not_echo_the_input(client):
    resp = client.post("/api/projects", json={
        "name": "secret-looking-value", "company": "Co",
        "start_year": 2026, "start_month": 13, "end_year": 2026, "end_month": 12,
    })
    assert resp.status_code == 422
    assert "secret-looking-value" not in resp.text
    assert resp.json()["detail"][0]["loc"] == ["body", "start_month"]


# ---------------------------------------------------------------------------
# F-41: uncounted register rows are named
# ---------------------------------------------------------------------------

def test_uncounted_reasons():
    d = date
    assert dep.uncounted_reason("", None, None, 100) == "no purchase type"
    assert dep.uncounted_reason("Rental", d(2026, 1, 1), None, 100) == \
        "unknown purchase type 'Rental'"
    assert dep.uncounted_reason("Not Purchased", None, None, 100) is None
    assert dep.uncounted_reason("Purchase", None, None, 0) is None
    assert dep.uncounted_reason("Purchase", None, None, 100) == "no purchase date"
    assert dep.uncounted_reason("Purchase", d(225, 1, 1), None, 100) == \
        "purchase date outside 1990-2100"
    assert dep.uncounted_reason("Purchase", d(2026, 1, 1), None, 100) is None
    assert dep.uncounted_reason("Leasing", d(2026, 1, 1), None, 100) == "no end date"
    assert dep.uncounted_reason("Leasing", d(2026, 1, 1), d(2205, 1, 1), 100) == \
        "end date outside 1990-2100"
    assert dep.uncounted_reason("Leasing", d(2026, 1, 1), d(2025, 1, 1), 100) == \
        "end date before purchase date"
    assert dep.uncounted_reason("Leasing", d(2026, 1, 1), d(2029, 1, 1), 100) is None
    assert dep.uncounted_reason("Planned Purchase", d(2026, 1, 1), None, 100) is None


def test_registers_and_summaries_report_uncounted_rows(client):
    hw_id = make_project(client, "Uncounted")
    resp = client.post(f"/api/hw/projects/{hw_id}/assets", json=asset(
        name="Undated", purchase_type="Purchase", purchase_cost=500.0,
    ))
    assert resp.status_code == 201, resp.text
    assert resp.json()["uncounted_reason"] == "no purchase date"
    assert resp.json()["total"] == 0.0
    resp = client.post(f"/api/hw/projects/{hw_id}/licenses", json=hw_license(
        name="Open-ended lease", depreciation="Leasing", purchase_cost=500.0,
        purchase_date="2026-01-01",
    ))
    assert resp.status_code == 201, resp.text
    assert resp.json()["uncounted_reason"] == "no end date"
    resp = client.post(f"/api/hw/projects/{hw_id}/licenses", json=hw_license(
        name="Counts", depreciation="Purchase", purchase_cost=500.0,
        purchase_date="2026-01-01",
    ))
    assert resp.json()["uncounted_reason"] is None

    summary = client.get(f"/api/hw/projects/{hw_id}/summary").json()
    assert summary["uncounted_rows"] == 2
    overview = client.get("/api/hw/overview").json()
    assert overview["uncounted_rows"] >= 2


# ---------------------------------------------------------------------------
# F-42: fixed roles carry no periods
# ---------------------------------------------------------------------------

def test_fixed_roles_drop_stray_periods_and_export_round_trips(client):
    project = new_project(client, name="Fixed export")
    pid = project["id"]
    feature_id = new_feature(client, pid)
    role = new_role(client, feature_id, ftes=1.0, use_advanced_allocation=False,
                    allocations=[{"start_month": "2026-01", "end_month": "2026-03",
                                  "ftes": 3.0}])
    assert role["allocations"] == []

    # A variable role switched back to fixed loses its periods too
    variable = new_role(client, feature_id, name="Tester", ftes=0.0,
                        use_advanced_allocation=True,
                        allocations=[{"start_month": "2026-01", "end_month": "2026-12",
                                      "ftes": 2.5}])
    resp = client.put(f"/api/roles/{variable['id']}", json={
        "name": "Tester", "location": "BCC", "level": "Senior", "ftes": 1.0,
        "use_advanced_allocation": False,
        "allocations": [{"start_month": "2026-01", "end_month": "2026-12", "ftes": 2.5}],
    })
    assert resp.status_code == 200, resp.text
    assert resp.json()["allocations"] == []

    exported = client.get(f"/api/projects/{pid}/export").json()
    for feature in exported["features"]:
        for exported_role in feature["Roles"]:
            assert exported_role["allocations"] == []
    resp = client.post("/api/projects/import", json=exported)
    assert resp.status_code == 201, resp.text


# ---------------------------------------------------------------------------
# F-43: importer edge cases
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("raw,expected", [
    ("0,500", 0.5), ("0.500", 0.5), (".5", 0.5), (",5", 0.5),
    ("1.234", 1234.0), ("1,234", 1234.0), ("1.234,50", 1234.5), ("1,234.50", 1234.5),
    ("29,90", 29.9), ("12.5", 12.5), ("€ 1.500", 1500.0), ("-7,25", -7.25),
    ("inf", None), ("nan", None), ("1e400", None), ("abc", None),
])
def test_parse_number_edges(raw, expected):
    assert hw_excel._parse_number(raw) == expected


def test_import_rejects_a_sheet_without_a_header_and_flags_alias_sheets(client):
    hw_id = make_project(client, "Headers")
    headerless = build_workbook([("Assets", ["Foo", "Bar"], [["x", "y"]])])
    resp = upload(client, hw_id, headerless, dry_run=True)
    assert resp.status_code == 400
    assert "no header row found" in resp.json()["detail"]

    two_assets = build_workbook([
        ("Assets", hw_excel.ASSET_HEADERS, REGISTER_ROWS),
        ("Asset", hw_excel.ASSET_HEADERS, REGISTER_ROWS),
    ])
    preview = upload(client, hw_id, two_assets, dry_run=True).json()
    assert len(preview["assets"]) == 1
    assert any("Several sheets look like the Assets register" in w
               for w in preview["warnings"])


def test_import_skips_a_hand_written_total_line_and_caps_absurd_amounts(client):
    hw_id = make_project(client, "Footer")
    workbook = build_workbook([("Assets", hw_excel.ASSET_HEADERS, REGISTER_ROWS + [
        ["", "", "Total", "", "", "", "", "", "", 1000, "", "", "", "", "", ""],
        ["A-9", "Co", "Too rich", "", "", "", "", "", "2026-01-01", 1e15, "", "", "", "",
         "", "Purchase"],
    ])])
    preview = upload(client, hw_id, workbook, dry_run=True).json()
    names = [row["name"] for row in preview["assets"]]
    assert "Total" not in names
    rich = next(row for row in preview["assets"] if row["name"] == "Too rich")
    assert rich["purchase_cost"] == 0.0
    assert any("larger than any real amount" in w for w in preview["warnings"])


# ---------------------------------------------------------------------------
# F-44: money is stored to the cent and totals are rounded once
# ---------------------------------------------------------------------------

def test_money_is_rounded_to_cents_on_the_way_in(client):
    hw_id = make_project(client, "Cents", budget_assets=1234.5678)
    assert client.get(f"/api/hw/projects/{hw_id}").json()["budget_assets"] == 1234.57
    resp = client.post(f"/api/hw/projects/{hw_id}/assets", json=asset(
        name="Third", purchase_type="Purchase", purchase_cost=1000 / 3,
        purchase_date="2026-01-01",
    ))
    assert resp.json()["purchase_cost"] == 333.33


def test_register_totals_are_rounded_once_from_full_precision():
    costs = dep.per_year("Leasing", date(2025, 7, 2), date(2028, 7, 2), 7157.35,
                         [2025, 2026, 2027, 2028])
    assert costs == {"2025": 1192.89, "2026": 2385.78, "2027": 2385.78, "2028": 1391.71}
    detailed = dep.license_year_costs(
        type("Row", (), {"depreciation": "Leasing", "purchase_date": date(2025, 7, 2),
                         "termination_date": date(2028, 7, 2), "purchase_cost": 7157.35})(),
        [2025, 2026, 2027, 2028],
    )
    assert detailed["total"] == round(sum(detailed["raw"]), 2) == 7356.17
    assert round(sum(detailed["per_year"].values()), 2) == 7356.16


# ---------------------------------------------------------------------------
# F-45: uniform grids and templates above the fixed cap
# ---------------------------------------------------------------------------

def test_uniform_grid_above_two_ftes_becomes_one_variable_period(client):
    project = new_project(client, start=(2026, 1), end=(2026, 3))
    pid = project["id"]
    feature_id = new_feature(client, pid)
    role = new_role(client, feature_id)
    resp = client.put(f"/api/projects/{pid}/resource-grid", json={"roles": [
        {"role_id": role["id"],
         "ftes_by_month": {"2026-01": 4.0, "2026-02": 4.0, "2026-03": 4.0}},
    ]})
    assert resp.status_code == 200, resp.text
    saved = next(r for f in client.get(f"/api/projects/{pid}").json()["features"]
                 for r in f["roles"] if r["id"] == role["id"])
    assert saved["use_advanced_allocation"] is True
    assert saved["ftes"] == 0.0
    assert [(p["start_month"], p["end_month"], p["ftes"]) for p in saved["allocations"]] == [
        ("2026-01", "2026-03", 4.0),
    ]

    # ... and back down to a fixed role when the grid is uniform and small
    resp = client.put(f"/api/projects/{pid}/resource-grid", json={"roles": [
        {"role_id": role["id"],
         "ftes_by_month": {"2026-01": 1.5, "2026-02": 1.5, "2026-03": 1.5}},
    ]})
    assert resp.status_code == 200, resp.text
    saved = next(r for f in client.get(f"/api/projects/{pid}").json()["features"]
                 for r in f["roles"] if r["id"] == role["id"])
    assert saved["use_advanced_allocation"] is False
    assert saved["ftes"] == 1.5
    assert saved["allocations"] == []


def test_template_roles_above_two_ftes_become_variable(client):
    source = new_project(client, name="Template source", start=(2026, 1), end=(2026, 6))
    feature_id = new_feature(client, source["id"], "Platform")
    new_role(client, feature_id, name="Crew", ftes=0.0, use_advanced_allocation=True,
             allocations=[{"start_month": "2026-01", "end_month": "2026-06", "ftes": 4.0}])
    resp = client.post(f"/api/projects/{source['id']}/save-as-template",
                       json={"name": "Big crew"})
    assert resp.status_code == 201, resp.text
    template_id = resp.json()["id"]

    created = new_project(client, name="From template", start=(2027, 1), end=(2027, 12),
                          template_id=template_id)
    role = created["features"][0]["roles"][0]
    assert role["use_advanced_allocation"] is True
    assert role["ftes"] == 0.0
    assert [(p["start_month"], p["end_month"], p["ftes"]) for p in role["allocations"]] == [
        ("2027-01", "2027-12", 4.0),
    ]
    assert client.get(f"/api/projects/{created['id']}/validate").json()["valid"] is True


# ---------------------------------------------------------------------------
# F-47: clones keep the lost reason, exports carry the hardware plan
# ---------------------------------------------------------------------------

def test_clone_keeps_the_lost_reason_and_export_carries_hardware(client):
    project = new_project(client, name="Lost one", status="lost",
                          lost_reason="Price", start=(2026, 1), end=(2026, 12))
    pid = project["id"]
    resp = client.post(f"/api/projects/{pid}/clone", json={"name": "Copy", "as_scenario": False})
    assert resp.status_code == 201, resp.text
    assert resp.json()["lost_reason"] == "Price"

    resp = client.post("/api/hardware-catalog", json={
        "name": "CANoe", "unit_cost": 10.0, "supplier_email": "sales@vector.example",
    })
    catalog_id = resp.json()["id"]
    resp = client.post(f"/api/projects/{pid}/hardware", json={
        "name": "CANoe", "catalog_item_id": catalog_id, "billing": "yearly",
        "unit_cost": 10.0, "qty": 3, "years": [2026], "supplier_name": "Vector",
    })
    assert resp.status_code == 201, resp.text

    exported = client.get(f"/api/projects/{pid}/export").json()
    assert exported["hardware_items"] == [{
        "name": "CANoe", "aspice": "SWE.3", "billing": "yearly", "unit_cost": 10.0,
        "qty": 3, "years": [2026], "supplier_name": "Vector", "supplier_email": "",
        "catalog_item_name": "CANoe",
    }]

    resp = client.post("/api/projects/import", json=exported)
    assert resp.status_code == 201, resp.text
    plan = client.get(f"/api/projects/{resp.json()['id']}/hardware").json()
    assert len(plan["items"]) == 1
    assert plan["items"][0]["catalog_item_id"] == catalog_id
    assert plan["items"][0]["supplier_email"] == "sales@vector.example"
    assert plan["grand_total"] == 30.0

    exported["hardware_items"][0]["years"] = [2031]
    resp = client.post("/api/projects/import", json=exported)
    assert resp.status_code == 422
    assert "outside the project years" in resp.text


def test_hardware_plan_warns_about_rows_outside_the_timeline():
    db = SessionLocal()
    try:
        project = models.Project(name="Warned", company="Co", start_year=2026,
                                 start_month=1, end_year=2026, end_month=12)
        db.add(project)
        db.flush()
        db.add(models.HardwareItem(project_id=project.id, name="Late rig", unit_cost=5.0,
                                   qty=1, billing="yearly", years_json="[2026, 2030]"))
        db.commit()
        db.refresh(project)
        from app.routers.hardware import build_plan
        plan = build_plan(project)
        assert plan["per_year"] == {2026: 5.0}
        assert plan["grand_total"] == 10.0
        assert plan["warnings"] == [
            "Late rig is planned for 2030, outside the project years 2026-2026, "
            "and is not shown in a year column"
        ]
        db.delete(project)
        db.commit()
    finally:
        db.close()


# ---------------------------------------------------------------------------
# F-35: an empty status filter means no status
# ---------------------------------------------------------------------------

def test_capacity_with_an_empty_status_filter_is_empty(client):
    project = new_project(client, name="Capacity", status="won")
    feature_id = new_feature(client, project["id"])
    new_role(client, feature_id)

    everything = client.get("/api/portfolio/capacity").json()
    assert everything["project_count"] >= 1
    won = client.get("/api/portfolio/capacity", params={"statuses": "won"}).json()
    assert won["project_count"] >= 1
    nothing = client.get("/api/portfolio/capacity", params={"statuses": ""}).json()
    assert nothing["project_count"] == 0
    assert nothing["cells"] == {}


# ---------------------------------------------------------------------------
# F-26: long-expired licenses leave the renewal list
# ---------------------------------------------------------------------------

def test_expiring_list_has_a_lower_bound(client):
    hw_id = make_project(client, "Expiry bound")
    today = date.today()
    for name, expires in (
        ("Ancient", today - timedelta(days=400)),
        ("Recent", today - timedelta(days=30)),
        ("Soon", today + timedelta(days=30)),
    ):
        resp = client.post(f"/api/hw/projects/{hw_id}/licenses", json=hw_license(
            name=name, expiration_date=expires.isoformat(),
        ))
        assert resp.status_code == 201, resp.text
    summary = client.get(f"/api/hw/projects/{hw_id}/summary").json()
    assert [row["name"] for row in summary["expiring"]] == ["Recent", "Soon"]
    # The counters still know about the old one
    assert summary["risk"]["expired"] == 2
