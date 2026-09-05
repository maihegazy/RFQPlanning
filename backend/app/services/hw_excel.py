"""Excel I/O for the Hardware Management module.

Writes the workbook that replaces `HW_purchasing_working_document_V5.xlsx`
(Dashboard, Summary, Assets, Licenses, HW Catalogue) with XlsxWriter, and reads
the Assets/Licenses registers back with openpyxl.

The reader is deliberately forgiving: the files purchasing actually hands around
carry merged banner cells, `#REF!` leftovers, German decimal strings, currency
symbols and a stray leading space in the Licenses `" Name"` header.
"""

import datetime
import io
import math
import re
from typing import Any

import openpyxl
import xlsxwriter

from ..config import (
    DATE_WINDOW_YEARS,
    HW_ASSET_CATEGORIES,
    HW_ASSET_STATUSES,
    HW_LICENSE_CATEGORIES,
    HW_PURCHASE_TYPES,
    MAX_MONEY,
    MAX_QUANTITY,
)
from .hw_depreciation import asset_year_costs, license_year_costs

FIRST_YEAR, LAST_YEAR = DATE_WINDOW_YEARS

# User text must never become a formula in the recipient's Excel.
WORKBOOK_OPTIONS = {"in_memory": True, "strings_to_formulas": False}

# A hand-written total line, in the ID column (our own footer) or under the name.
FOOTER_WORDS = ("total", "grand total")

# Palette and number formats shared with the budget workbook
# (frontend/src/money/excelBudget.ts) so both exports look like one product.
YELLOW = "#FFFF00"
GRAY = "#D3D3D3"
GREEN = "#90EE90"
PINK = "#FFB6C1"
DARK = "#1F2937"
LIGHT_GRAY = "#F3F4F6"

EURO_FMT = '#,##0.00 "€"'
COUNT_FMT = "#,##0"
DATE_FMT = "yyyy-mm-dd"

DEFAULT_PURCHASE_TYPE = "Not Purchased"

# (header, model field, kind, blank default) in sheet column order.
ASSET_FIELDS: list[tuple[str, str, str, Any]] = [
    ("ID", "asset_tag", "text", ""),
    ("Company", "company", "text", ""),
    ("Asset Name", "name", "text", ""),
    ("Serial", "serial", "text", ""),
    ("Model", "model", "text", ""),
    ("Category", "category", "text", ""),
    ("Status", "status", "text", ""),
    ("Supplier", "supplier", "text", ""),
    ("Purchase Date", "purchase_date", "date", None),
    ("Purchase Cost", "purchase_cost", "money", 0.0),
    ("Order Number", "order_number", "text", ""),
    ("EOL Date", "eol_date", "date", None),
    ("Assigned Employee", "assigned_employee", "text", ""),
    ("SW License", "sw_license", "text", ""),
    ("Purchased by", "purchased_by", "text", ""),
    ("Purchase Type", "purchase_type", "choice", DEFAULT_PURCHASE_TYPE),
]

LICENSE_FIELDS: list[tuple[str, str, str, Any]] = [
    ("ID", "license_tag", "text", ""),
    ("Company", "company", "text", ""),
    ("Name", "name", "text", ""),
    ("Product Key", "product_key", "text", ""),
    ("Expiration Date", "expiration_date", "date", None),
    ("Licensed to Email", "licensed_to_email", "text", ""),
    ("Category", "category", "text", ""),
    ("Supplier", "supplier", "text", ""),
    ("Manufacturer", "manufacturer", "text", ""),
    ("Total", "quantity", "int", 1),
    ("Purchase Date", "purchase_date", "date", None),
    ("Termination Date", "termination_date", "date", None),
    ("Depreciation", "depreciation", "choice", DEFAULT_PURCHASE_TYPE),
    ("Maintained", "maintained", "bool", False),
    ("Purchase Cost", "purchase_cost", "money", 0.0),
    ("Purchase Order Number", "purchase_order_number", "text", ""),
    ("Notes", "notes", "text", ""),
]

ASSET_HEADERS = [field[0] for field in ASSET_FIELDS]
LICENSE_HEADERS = [field[0] for field in LICENSE_FIELDS]

CATALOG_HEADERS = ["Supplier", "Asset", "Type", "Price (Euro)", "Billing", "ASPICE",
                   "Contact"]

# Summary!A4:J4, kept in the original column order so the sheet still reads the
# way the purchasing manager's spreadsheet did.
SUMMARY_HEADERS = ["Year", "Total Actual + Planned", "Total Actual Budget", "HW Budget",
                   "Licenses Budget", "Total Planned Budget", "Planned Budget - HW",
                   "Planned Budget - Licenses", "Special Cases Budget - Assets",
                   "Special Cases Budget - Licenses"]
SUMMARY_KEYS = ["grand_total", "actual_total", "actual_assets", "actual_licenses",
                "planned_total", "planned_assets", "planned_licenses"]

ASSET_SHEET = "Assets"
LICENSE_SHEET = "Licenses"

# Sheet-title spellings accepted on import, normalised.
SHEET_ALIASES = {
    ASSET_SHEET: {"assets", "asset", "assets register", "hardware assets"},
    LICENSE_SHEET: {"licenses", "license", "licences", "licence", "licenses register"},
}

# Formula leftovers of the original workbook read back as these strings.
ERROR_VALUES = {"#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!",
                "#SPILL!", "#CALC!", "#GETTING_DATA"}

YEAR_RE = re.compile(r"^\d{4}$")
NUMBER_JUNK_RE = re.compile(r"[€$£']|\bEUR\b|\bEURO\b", re.IGNORECASE)
WHITESPACE_RE = re.compile(r"\s")  # \s covers the NBSP Excel likes to paste

DATE_FORMATS = ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"]
EXCEL_EPOCH = datetime.date(1899, 12, 30)  # Excel's 1900 leap-year bug included
EXCEL_SERIAL_MAX = 2958465  # 9999-12-31

TRUE_WORDS = {"true", "yes", "y", "x", "1", "ja", "wahr"}
FALSE_WORDS = {"false", "no", "n", "0", "nein", "falsch", ""}

# A banner or title block above the real header row is common in hand-edited files.
HEADER_SCAN_ROWS = 10
MIN_HEADER_MATCHES = 2

# One merged range can span a whole column; only expand ranges small enough to be
# a real merged label.
MAX_MERGE_CELLS = 4096

MAX_WARNINGS = 200

# Column lengths of the register models, so an over-long cell is cut here with a
# warning instead of failing the whole upload in request validation.
DEFAULT_TEXT_LIMIT = 255
TEXT_LIMITS = {"notes": 4000}

TEMPLATE_VALIDATION_ROWS = 500


# --------------------------------------------------------------------------- #
# shared helpers
# --------------------------------------------------------------------------- #

def _sanitize_sheet_name(name: str) -> str:
    for char in [':', '\\', '/', '?', '*', '[', ']']:
        name = name.replace(char, '_')
    return name[:31]


def _norm(value: Any) -> str:
    """Lower-cased, whitespace-collapsed key used to match headers and words."""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return " ".join(str(value if value is not None else "").split()).lower()


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == "" or value.strip() in ERROR_VALUES
    return False


def _text(value: Any) -> str:
    if _is_blank(value):
        return ""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _as_date(value: Any) -> datetime.date | None:
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return None


def _money(value: Any) -> float:
    try:
        return round(float(value or 0.0), 2)
    except (TypeError, ValueError):
        return 0.0


# --------------------------------------------------------------------------- #
# writing
# --------------------------------------------------------------------------- #

class _Widths:
    """Column widths grown from the content written into them."""

    def __init__(self, minimum: int = 9, cap: int = 44):
        self._widths: dict[int, int] = {}
        self._min = minimum
        self._cap = cap

    def note(self, col: int, value: Any) -> None:
        longest = max((len(part) for part in _text(value).split("\n")), default=0)
        self._widths[col] = max(self._widths.get(col, self._min), longest + 2)

    def apply(self, worksheet) -> None:
        for col, width in self._widths.items():
            worksheet.set_column(col, col, min(width, self._cap))


def _formats(workbook) -> dict[str, Any]:
    base = {'border': 1, 'valign': 'vcenter'}
    return {
        "title": workbook.add_format({
            'bold': True, 'font_size': 14, 'font_color': 'white', 'bg_color': DARK,
            'align': 'left', 'valign': 'vcenter',
        }),
        "subtitle": workbook.add_format({'italic': True, 'font_color': '#6B7280'}),
        "section": workbook.add_format({
            'bold': True, 'font_color': 'white', 'bg_color': DARK, 'align': 'left',
            **base,
        }),
        "header": workbook.add_format({
            'bold': True, 'bg_color': YELLOW, 'font_color': 'black', 'align': 'center',
            'text_wrap': True, **base,
        }),
        "label": workbook.add_format({'bold': True, 'bg_color': LIGHT_GRAY,
                                      'align': 'left', **base}),
        "text": workbook.add_format({'align': 'left', **base}),
        "wrap": workbook.add_format({'align': 'left', 'text_wrap': True, **base}),
        "money": workbook.add_format({'num_format': EURO_FMT, **base}),
        "money_total": workbook.add_format({'bold': True, 'bg_color': GRAY,
                                            'num_format': EURO_FMT, **base}),
        "money_good": workbook.add_format({'bold': True, 'bg_color': GREEN,
                                           'num_format': EURO_FMT, **base}),
        "money_bad": workbook.add_format({'bold': True, 'bg_color': PINK,
                                          'num_format': EURO_FMT, **base}),
        "count": workbook.add_format({'num_format': COUNT_FMT, 'align': 'center',
                                      **base}),
        "count_total": workbook.add_format({'bold': True, 'bg_color': GRAY,
                                            'num_format': COUNT_FMT,
                                            'align': 'center', **base}),
        "date": workbook.add_format({'num_format': DATE_FMT, 'align': 'center', **base}),
        "bool": workbook.add_format({'align': 'center', **base}),
        "total_label": workbook.add_format({'bold': True, 'bg_color': GRAY,
                                            'align': 'left', **base}),
    }


def _write_cell(worksheet, row: int, col: int, value: Any, kind: str,
                fmts: dict[str, Any], widths: _Widths) -> None:
    """Write one register cell in the format its column kind calls for."""
    if kind == "date":
        parsed = _as_date(value)
        if parsed is None:
            worksheet.write_blank(row, col, None, fmts["date"])
        else:
            worksheet.write_datetime(row, col, parsed, fmts["date"])
        widths.note(col, "0000-00-00")
    elif kind == "money":
        worksheet.write_number(row, col, _money(value), fmts["money"])
        widths.note(col, f"{_money(value):,.2f} €")
    elif kind == "int":
        worksheet.write_number(row, col, int(value or 0), fmts["count"])
        widths.note(col, value)
    elif kind == "bool":
        worksheet.write_boolean(row, col, bool(value), fmts["bool"])
        widths.note(col, "FALSE")
    else:
        worksheet.write_string(row, col, _text(value), fmts["text"])
        widths.note(col, value)


def _write_register(workbook, fmts: dict[str, Any], sheet_name: str,
                    fields: list[tuple[str, str, str, Any]], rows: list[Any],
                    years: list[int], year_costs) -> None:
    """One register sheet: template headers, the rows, the derived year columns."""
    worksheet = workbook.add_worksheet(_sanitize_sheet_name(sheet_name))
    widths = _Widths()

    for col, (header, _field, _kind, _default) in enumerate(fields):
        worksheet.write_string(0, col, header, fmts["header"])
        widths.note(col, header)
    year_start = len(fields)
    for offset, year in enumerate(years):
        worksheet.write_string(0, year_start + offset, str(year), fmts["header"])
        widths.note(year_start + offset, "0,000,000.00 €")

    year_totals = [0.0] * len(years)
    for index, row in enumerate(rows, start=1):
        for col, (_header, field, kind, default) in enumerate(fields):
            _write_cell(worksheet, index, col,
                        getattr(row, field, default), kind, fmts, widths)
        costs = year_costs(row, years)
        for offset, year in enumerate(years):
            # Cells show cents; the footer sums full precision like the Summary sheet.
            year_totals[offset] += costs["raw"][offset]
            worksheet.write_number(index, year_start + offset,
                                   costs["per_year"][str(year)], fmts["money"])

    total_row = len(rows) + 1
    worksheet.write_string(total_row, 0, "TOTAL", fmts["total_label"])
    for col in range(1, year_start):
        kind = fields[col][2]
        if kind == "money":
            worksheet.write_number(
                total_row, col,
                round(sum(_money(getattr(r, fields[col][1], 0.0)) for r in rows), 2),
                fmts["money_total"],
            )
        else:
            worksheet.write_blank(total_row, col, None, fmts["total_label"])
    for offset in range(len(years)):
        worksheet.write_number(total_row, year_start + offset,
                               round(year_totals[offset], 2), fmts["money_total"])

    last_col = year_start + len(years) - 1
    worksheet.autofilter(0, 0, len(rows), last_col)
    worksheet.freeze_panes(1, 3)
    widths.apply(worksheet)


def _write_dashboard(workbook, fmts: dict[str, Any], project: Any, summary: dict,
                     today: datetime.date) -> None:
    worksheet = workbook.add_worksheet("Dashboard")
    dashboard = summary.get("dashboard", {})
    risk = summary.get("risk", {})

    worksheet.set_column(0, 0, 3)
    worksheet.set_column(1, 1, 46)
    worksheet.set_column(2, 2, 22)
    worksheet.set_column(3, 4, 18)

    worksheet.merge_range(0, 1, 0, 4,
                          f"Hardware Management — {_text(getattr(project, 'name', ''))}",
                          fmts["title"])
    worksheet.write_string(1, 1, f"Generated {today.isoformat()}", fmts["subtitle"])

    def section(row: int, title: str) -> int:
        worksheet.merge_range(row, 1, row, 4, title, fmts["section"])
        return row + 1

    def entry(row: int, label: str, value: Any, fmt: str, wide: bool = False) -> int:
        worksheet.write_string(row, 1, label, fmts["label"])
        if wide:
            worksheet.merge_range(row, 2, row, 4, _text(value), fmts[fmt])
        elif fmt.startswith("money"):
            worksheet.write_number(row, 2, _money(value), fmts[fmt])
        elif fmt == "count":
            worksheet.write_number(row, 2, int(value or 0), fmts[fmt])
        else:
            worksheet.write_string(row, 2, _text(value), fmts[fmt])
        return row + 1

    row = section(3, "Project")
    row = entry(row, "Company", getattr(project, "company", ""), "text", wide=True)
    row = entry(row, "Portal reference", getattr(project, "portal_reference", ""),
                "text", wide=True)
    row = entry(row, "Description", getattr(project, "description", ""), "wrap",
                wide=True)

    row = section(row + 1, "Budget")
    row = entry(row, "Overall Project Budget", dashboard.get("budget_total"), "money")
    row = entry(row, "Assets Budget", dashboard.get("budget_assets"), "money")
    row = entry(row, "Licenses Budget", dashboard.get("budget_licenses"), "money")
    row = entry(row, "Total Spent Budget (including confirmed depreciations)",
                dashboard.get("spent_total"), "money")
    row = entry(row, "Planned Budget", dashboard.get("planned_total"), "money")
    remaining = _money(dashboard.get("remaining"))
    row = entry(row, "Remaining Budget", remaining,
                "money_good" if remaining >= 0 else "money_bad")

    row = section(row + 1, "License Renewal Risk")
    row = entry(row, "Licenses expired", risk.get("expired"), "count")
    row = entry(row, "Licenses expiring in 30 days", risk.get("in_30_days"), "count")
    row = entry(row, "Licenses expiring in 60 days", risk.get("in_60_days"), "count")
    row = entry(row, "Licenses expiring in 90 days", risk.get("in_90_days"), "count")

    row = section(row + 1, "Register")
    row = entry(row, "Assets", summary.get("asset_count"), "count")
    row = entry(row, "Licenses", summary.get("license_count"), "count")

    expiring = summary.get("expiring") or []
    if expiring:
        row = section(row + 1, "Renewals: expired within the last year, or due within 90 days")
        for col, header in enumerate(["Name", "Manufacturer", "Expiration", "Days left"]):
            worksheet.write_string(row, col + 1, header, fmts["header"])
        row += 1
        for item in expiring:
            worksheet.write_string(row, 1, _text(item.get("name")), fmts["text"])
            worksheet.write_string(row, 2, _text(item.get("manufacturer")), fmts["text"])
            worksheet.write_string(row, 3, _text(item.get("expiration_date")),
                                   fmts["text"])
            worksheet.write_number(row, 4, int(item.get("days_left") or 0),
                                   fmts["count"])
            row += 1


def _write_pivot_block(worksheet, fmts: dict[str, Any], widths: _Widths, row: int,
                       title: str, row_label: str, pivot: dict) -> int:
    """Summary r26-r37: counts of rows by category down, status across."""
    statuses = pivot.get("statuses", [])
    worksheet.write_string(row, 0, title, fmts["label"])
    worksheet.write_string(row, 1, "Column Labels", fmts["label"])
    widths.note(0, title)
    row += 1

    worksheet.write_string(row, 0, row_label, fmts["header"])
    for col, status in enumerate(statuses, start=1):
        worksheet.write_string(row, col, status, fmts["header"])
        widths.note(col, status)
    worksheet.write_string(row, len(statuses) + 1, "Total", fmts["header"])
    row += 1

    column_totals = [0] * len(statuses)
    for entry in pivot.get("rows", []):
        worksheet.write_string(row, 0, _text(entry.get("category")), fmts["text"])
        widths.note(0, entry.get("category"))
        counts = entry.get("counts", {})
        for col, status in enumerate(statuses, start=1):
            value = int(counts.get(status, 0) or 0)
            column_totals[col - 1] += value
            worksheet.write_number(row, col, value, fmts["count"])
        worksheet.write_number(row, len(statuses) + 1, int(entry.get("total", 0) or 0),
                               fmts["count_total"])
        row += 1

    worksheet.write_string(row, 0, "Grand Total", fmts["total_label"])
    for col, value in enumerate(column_totals, start=1):
        worksheet.write_number(row, col, value, fmts["count_total"])
    worksheet.write_number(row, len(statuses) + 1, sum(column_totals),
                           fmts["count_total"])
    return row + 1


def _write_summary(workbook, fmts: dict[str, Any], summary: dict) -> None:
    worksheet = workbook.add_worksheet("Summary")
    widths = _Widths()

    special: dict[tuple[int, str], float] = {}
    for adjustment in summary.get("adjustments", []):
        key = (int(adjustment.get("year") or 0), str(adjustment.get("kind") or ""))
        special[key] = special.get(key, 0.0) + _money(adjustment.get("amount"))

    for col, header in enumerate(SUMMARY_HEADERS):
        worksheet.write_string(0, col, header, fmts["header"])
        widths.note(col, header)

    row = 1
    for year_row in summary.get("years", []):
        year = int(year_row.get("year") or 0)
        worksheet.write_number(row, 0, year, fmts["count"])
        for col, key in enumerate(SUMMARY_KEYS, start=1):
            worksheet.write_number(row, col, _money(year_row.get(key)), fmts["money"])
        worksheet.write_number(row, 8, special.get((year, "assets"), 0.0), fmts["money"])
        worksheet.write_number(row, 9, special.get((year, "licenses"), 0.0),
                               fmts["money"])
        row += 1

    totals = summary.get("totals", {})
    worksheet.write_string(row, 0, "Total", fmts["total_label"])
    for col, key in enumerate(SUMMARY_KEYS, start=1):
        worksheet.write_number(row, col, _money(totals.get(key)), fmts["money_total"])
    worksheet.write_number(row, 8, round(sum(v for (_y, k), v in special.items()
                                             if k == "assets"), 2), fmts["money_total"])
    worksheet.write_number(row, 9, round(sum(v for (_y, k), v in special.items()
                                             if k == "licenses"), 2),
                           fmts["money_total"])
    for col in range(len(SUMMARY_HEADERS)):
        widths.note(col, "0,000,000.00 €")

    row += 3
    worksheet.write_string(row, 0, "Asset Type", fmts["header"])
    worksheet.write_string(row, 1, "Count", fmts["header"])
    worksheet.write_string(row + 1, 0, "Hardware", fmts["text"])
    worksheet.write_number(row + 1, 1, int(summary.get("asset_count", 0) or 0),
                           fmts["count"])
    worksheet.write_string(row + 2, 0, "License", fmts["text"])
    worksheet.write_number(row + 2, 1, int(summary.get("license_count", 0) or 0),
                           fmts["count"])

    row += 5
    row = _write_pivot_block(worksheet, fmts, widths, row, "Count of Asset Name",
                             "Row Labels", summary.get("asset_pivot", {}))
    row += 2
    _write_pivot_block(worksheet, fmts, widths, row, "Count of License Name",
                       "Row Labels", summary.get("license_pivot", {}))

    worksheet.freeze_panes(1, 1)
    widths.apply(worksheet)


def _write_catalogue(workbook, fmts: dict[str, Any], catalog_items: list[Any]) -> None:
    worksheet = workbook.add_worksheet("HW Catalogue")
    widths = _Widths()

    for col, header in enumerate(CATALOG_HEADERS):
        worksheet.write_string(0, col, header, fmts["header"])
        widths.note(col, header)

    items = sorted(
        catalog_items,
        key=lambda item: (_norm(getattr(item, "supplier_name", "")),
                          _norm(getattr(item, "name", ""))),
    )
    for index, item in enumerate(items, start=1):
        billing = _text(getattr(item, "billing", ""))
        # The original sheet's Type column separated H/W from License; the shared
        # catalogue only records that split through the billing mode (a yearly
        # item is a licence/subscription, a one-off purchase is hardware).
        kind = "License" if _norm(billing) == "yearly" else "H/W"
        values = [
            _text(getattr(item, "supplier_name", "")),
            _text(getattr(item, "name", "")),
            kind,
            None,
            billing,
            _text(getattr(item, "aspice", "")),
            _text(getattr(item, "supplier_email", "")),
        ]
        for col, value in enumerate(values):
            if col == 3:
                worksheet.write_number(index, col,
                                       _money(getattr(item, "unit_cost", 0.0)),
                                       fmts["money"])
                widths.note(col, "0,000.00 €")
            else:
                worksheet.write_string(index, col, value, fmts["text"])
                widths.note(col, value)

    worksheet.autofilter(0, 0, len(items), len(CATALOG_HEADERS) - 1)
    worksheet.freeze_panes(1, 0)
    widths.apply(worksheet)


def build_project_workbook(project: Any, assets: list[Any], licenses: list[Any],
                           summary: dict, catalog_items: list[Any],
                           today: datetime.date) -> bytes:
    """The full working document for one HW project, as xlsx bytes."""
    years = [int(row["year"]) for row in summary.get("years", [])]

    buffer = io.BytesIO()
    workbook = xlsxwriter.Workbook(buffer, WORKBOOK_OPTIONS)
    fmts = _formats(workbook)

    _write_dashboard(workbook, fmts, project, summary, today)
    _write_summary(workbook, fmts, summary)
    _write_register(workbook, fmts, ASSET_SHEET, ASSET_FIELDS, list(assets), years,
                    asset_year_costs)
    _write_register(workbook, fmts, LICENSE_SHEET, LICENSE_FIELDS, list(licenses),
                    years, license_year_costs)
    _write_catalogue(workbook, fmts, list(catalog_items))

    workbook.close()
    return buffer.getvalue()


def _write_template_sheet(workbook, fmts: dict[str, Any], sheet_name: str,
                          fields: list[tuple[str, str, str, Any]],
                          validations: dict[str, list[str]]) -> None:
    worksheet = workbook.add_worksheet(sheet_name)
    widths = _Widths()
    last_row = TEMPLATE_VALIDATION_ROWS
    date_columns = []

    for col, (header, _field, kind, _default) in enumerate(fields):
        worksheet.write_string(0, col, header, fmts["header"])
        widths.note(col, header)
        if kind == "date":
            date_columns.append(col)
        choices = validations.get(header)
        if choices:
            worksheet.data_validation(1, col, last_row, col, {
                'validate': 'list', 'source': choices,
                'error_message': f"Pick one of: {', '.join(choices)}",
                'error_type': 'warning',  # free text stays allowed, as in the sheet
            })

    worksheet.freeze_panes(1, 0)
    widths.apply(worksheet)
    for col in date_columns:  # after apply(): set_column keeps only the last call
        worksheet.set_column(col, col, 14, fmts["date"])


def _write_readme(workbook, fmts: dict[str, Any]) -> None:
    worksheet = workbook.add_worksheet("Read me")
    worksheet.set_column(0, 0, 26)
    worksheet.set_column(1, 1, 96)

    worksheet.merge_range(0, 0, 0, 1, "Hardware Management — import template",
                          fmts["title"])

    lines = [
        ("Sheets", f"Fill in '{ASSET_SHEET}' and/or '{LICENSE_SHEET}'. "
                   "Both are optional; other sheets are ignored."),
        ("Rows", "One row per asset / license. A row without a name is skipped."),
        ("Columns", "Matched by header name (case and spacing are ignored). "
                    "Unknown columns are reported and ignored. "
                    "Four-digit year columns are derived and never imported."),
        ("Dates", "Real Excel dates, or text as YYYY-MM-DD, DD.MM.YYYY or DD/MM/YYYY."),
        ("Numbers", "1234.56, 1.234,56, 1,234.56 or € 1234. Blank counts as 0."),
        ("Purchase Type", ", ".join(HW_PURCHASE_TYPES)
         + f" — anything else falls back to '{DEFAULT_PURCHASE_TYPE}'."),
        ("Depreciation", "Same values as Purchase Type; it is the license-side name "
                         "of the same field."),
        ("Maintained", "true / false / yes / no / x / 1 / 0. Blank counts as false."),
        ("Status", ", ".join(HW_ASSET_STATUSES) + " (free text is accepted too)."),
        ("Asset Category", ", ".join(HW_ASSET_CATEGORIES)
         + " (free text is accepted too)."),
        ("License Category", ", ".join(HW_LICENSE_CATEGORIES)
         + " (free text is accepted too)."),
        ("Total", "Licenses only: the quantity. Blank counts as 1."),
        ("Leasing", "Leasing rows are written off over 36 months from the purchase "
                    "date to the termination / EOL date."),
    ]
    for index, (label, text) in enumerate(lines, start=2):
        worksheet.write_string(index, 0, label, fmts["label"])
        worksheet.write_string(index, 1, text, fmts["wrap"])


def build_import_template() -> bytes:
    """Empty Assets + Licenses sheets carrying exactly the import headers."""
    buffer = io.BytesIO()
    workbook = xlsxwriter.Workbook(buffer, WORKBOOK_OPTIONS)
    fmts = _formats(workbook)

    _write_template_sheet(workbook, fmts, ASSET_SHEET, ASSET_FIELDS, {
        "Category": HW_ASSET_CATEGORIES,
        "Status": HW_ASSET_STATUSES,
        "Purchase Type": HW_PURCHASE_TYPES,
    })
    _write_template_sheet(workbook, fmts, LICENSE_SHEET, LICENSE_FIELDS, {
        "Category": HW_LICENSE_CATEGORIES,
        "Depreciation": HW_PURCHASE_TYPES,
        "Maintained": ["TRUE", "FALSE"],
    })
    _write_readme(workbook, fmts)

    workbook.close()
    return buffer.getvalue()


# --------------------------------------------------------------------------- #
# reading
# --------------------------------------------------------------------------- #

def _parse_number(value: Any) -> float | None:
    """German and English money strings; None when the text is not a number."""
    if _is_blank(value):
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, datetime.date):
        return None

    text = WHITESPACE_RE.sub("", NUMBER_JUNK_RE.sub("", str(value)))
    if not text:
        return 0.0
    sign = -1.0 if text.startswith("-") else 1.0
    text = text.lstrip("+-")

    dots, commas = text.count("."), text.count(",")
    if dots and commas:
        # Whichever separator comes last is the decimal point.
        decimal = "," if text.rfind(",") > text.rfind(".") else "."
        text = text.replace("," if decimal == "." else ".", "").replace(decimal, ".")
    elif dots or commas:
        separator = "." if dots else ","
        head, _, tail = text.rpartition(separator)
        # A single separator with a full group of three digits behind it is a
        # thousands separator in both conventions ("1.234" and "1,234" are 1234);
        # anything else is the decimal point ("29,90", "12.5"). A bare or zero
        # head can only be a decimal: ".5" and "0,500" are half, not five hundred.
        grouped = len(tail) == 3 and head.isdigit() and head != "0"
        if text.count(separator) > 1 or grouped:
            text = text.replace(separator, "")
        else:
            text = f"{head or '0'}.{tail}"
    try:
        value = sign * float(text)
    except ValueError:
        return None
    # "inf" and "nan" are floats to Python but not numbers to a register.
    return value if math.isfinite(value) else None


def _in_window(value: datetime.date) -> bool:
    return FIRST_YEAR <= value.year <= LAST_YEAR


def _parse_date(value: Any) -> datetime.date | None:
    """A date inside the window, or None (the caller warns when the cell was not blank)."""
    if _is_blank(value):
        return None
    parsed = _as_date(value)
    if parsed is not None:
        return parsed if _in_window(parsed) else None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # A date column that lost its number format reads back as a serial. A bare
        # year typed into the column is a number too (2026 would be 1905-07-18), so
        # only a serial that lands inside the window is trusted.
        if 1 <= value <= EXCEL_SERIAL_MAX:
            parsed = EXCEL_EPOCH + datetime.timedelta(days=int(value))
            return parsed if _in_window(parsed) else None
        return None

    text = str(value).strip().replace("T", " ").split(" ")[0]
    for fmt in DATE_FORMATS:
        try:
            parsed = datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
        return parsed if _in_window(parsed) else None
    return None


def _parse_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if _is_blank(value):
        return False
    if isinstance(value, (int, float)):
        return bool(value)
    word = _norm(value)
    if word in TRUE_WORDS:
        return True
    if word in FALSE_WORDS:
        return False
    return None


def _parse_choice(value: Any) -> str | None:
    if _is_blank(value):
        return DEFAULT_PURCHASE_TYPE
    wanted = _norm(value)
    for choice in HW_PURCHASE_TYPES:
        if _norm(choice) == wanted:
            return choice
    return None


def _cell_reader(worksheet):
    """Cell lookup that resolves merged ranges to their anchor value."""
    merged: dict[tuple[int, int], Any] = {}
    for cell_range in worksheet.merged_cells.ranges:
        rows = cell_range.max_row - cell_range.min_row + 1
        cols = cell_range.max_col - cell_range.min_col + 1
        if rows * cols > MAX_MERGE_CELLS:
            continue
        anchor = worksheet.cell(cell_range.min_row, cell_range.min_col).value
        if anchor is None:
            continue
        for row in range(cell_range.min_row, cell_range.max_row + 1):
            for col in range(cell_range.min_col, cell_range.max_col + 1):
                merged[(row, col)] = anchor

    def read(row: int, col: int) -> Any:
        value = worksheet.cell(row, col).value
        if value is None:
            value = merged.get((row, col))
        return None if _is_blank(value) else value

    return read


def _find_header_row(read, max_row: int, max_col: int, known: set[str]) -> int | None:
    """The row carrying the most known headers — files often start with a banner."""
    best_row, best_score = None, 0
    for row in range(1, min(max_row, HEADER_SCAN_ROWS) + 1):
        score = sum(1 for col in range(1, max_col + 1)
                    if _norm(read(row, col)) in known)
        if score > best_score:
            best_row, best_score = row, score
    return best_row if best_score >= MIN_HEADER_MATCHES else None


def _map_columns(read, header_row: int, max_col: int,
                 fields: list[tuple[str, str, str, Any]], sheet_name: str,
                 warnings: list[str]) -> dict[int, tuple[str, str, str, Any]]:
    by_header = {_norm(field[0]): field for field in fields}
    columns: dict[int, tuple[str, str, str, Any]] = {}
    used: set[str] = set()

    for col in range(1, max_col + 1):
        label = read(header_row, col)
        key = _norm(label)
        if not key:
            continue
        if YEAR_RE.match(key):
            continue  # year columns are derived from the depreciation engine
        field = by_header.get(key)
        if field is None:
            warnings.append(f"{sheet_name}: ignored unknown column '{_text(label)}'")
            continue
        if key in used:
            warnings.append(f"{sheet_name}: ignored duplicate column '{_text(label)}'")
            continue
        used.add(key)
        columns[col] = field
    return columns


def _parse_row(read, row: int, columns: dict[int, tuple[str, str, str, Any]],
               sheet_name: str, warnings: list[str]) -> dict[str, Any] | None:
    """One register row, or None when it holds nothing worth importing."""
    values: dict[str, Any] = {}
    has_data = False
    tag = ""
    name_header = "name"

    for col, (header, field, kind, default) in columns.items():
        raw = read(row, col)
        if kind == "date":
            parsed = _parse_date(raw)
            if parsed is None and not _is_blank(raw):
                warnings.append(f"{sheet_name} row {row}: '{header}' value "
                                f"'{_text(raw)}' is not a date between {FIRST_YEAR} "
                                f"and {LAST_YEAR}, left empty")
            values[field] = parsed.isoformat() if parsed else None
            has_data = has_data or parsed is not None
        elif kind in ("money", "int"):
            number = _parse_number(raw)
            limit = MAX_QUANTITY if kind == "int" else MAX_MONEY
            if number is None:
                warnings.append(f"{sheet_name} row {row}: '{header}' value "
                                f"'{_text(raw)}' is not a number, used {default}")
            elif number < 0:
                # Costs and quantities are `ge=0` on the API models; a negative
                # cell would fail the whole import instead of just this value.
                warnings.append(f"{sheet_name} row {row}: '{header}' value "
                                f"'{_text(raw)}' is negative, used 0")
                number = 0.0
            elif number > limit:
                warnings.append(f"{sheet_name} row {row}: '{header}' value "
                                f"'{_text(raw)}' is larger than any real "
                                f"{'quantity' if kind == 'int' else 'amount'}, "
                                f"used {default}")
                number = None
            if number is None or _is_blank(raw):
                values[field] = default
            else:
                values[field] = int(number) if kind == "int" else round(number, 2)
                has_data = has_data or number != 0
        elif kind == "bool":
            flag = _parse_bool(raw)
            if flag is None:
                warnings.append(f"{sheet_name} row {row}: '{header}' value "
                                f"'{_text(raw)}' is not a yes/no, used No")
                flag = False
            values[field] = flag
            has_data = has_data or flag
        elif kind == "choice":
            choice = _parse_choice(raw)
            if choice is None:
                warnings.append(f"{sheet_name} row {row}: unknown {header} "
                                f"'{_text(raw)}', using '{DEFAULT_PURCHASE_TYPE}'")
                choice = DEFAULT_PURCHASE_TYPE
            values[field] = choice
            # A dropdown never makes a row count as filled in: the real working
            # document ships hundreds of blank rows whose only content is a
            # pre-selected Purchase Type.
        else:
            text = _text(raw)
            limit = TEXT_LIMITS.get(field, DEFAULT_TEXT_LIMIT)
            if len(text) > limit:
                warnings.append(f"{sheet_name} row {row}: '{header}' was longer than "
                                f"{limit} characters and was cut short")
                text = text[:limit]
            values[field] = text
            has_data = has_data or bool(text)
            if field in ("asset_tag", "license_tag"):
                tag = text
            elif field == "name":
                name_header = header

    if (
        _norm(values.get("name")) in FOOTER_WORDS
        and _norm(tag) in ("", *FOOTER_WORDS)
        and not values.get("serial")
        and not values.get("purchase_date")
    ):
        return None  # a total line written under the name column by hand

    if not values.get("name"):
        # "TOTAL" is the footer our own export writes, so re-importing an export
        # does not report it as a broken row.
        if not has_data or _norm(tag) in FOOTER_WORDS:
            return None
        # The real working document leaves the name blank on most licence rows and
        # identifies them by category + manufacturer instead. Dropping those rows
        # would throw away the whole register, so name them from what is there and
        # let the user rename them afterwards. A row carrying nothing that could
        # identify it is still skipped — there is nothing to name it after.
        synthesised = _synthesise_name(values)
        if synthesised is None:
            warnings.append(f"{sheet_name} row {row}: skipped, no {name_header}")
            return None
        values["name"] = synthesised
        warnings.append(f"{sheet_name} row {row}: no {name_header}, "
                        f"named '{synthesised}' from the other columns")
    values["catalog_item_id"] = None
    return values


def _synthesise_name(values: dict[str, Any]) -> str | None:
    """Name a row that has data but no name of its own, or None if nothing fits.

    Only genuinely identifying columns are used: a row carrying nothing but, say,
    a company is not something a synthesised name would help anyone find.
    """
    if "license_tag" in values:
        parts = [values.get("category"), values.get("manufacturer")]
        fallbacks = ("supplier", "product_key", "license_tag")
    else:
        parts = [values.get("model"), values.get("category")]
        fallbacks = ("supplier", "serial", "asset_tag")

    named = " — ".join(str(p).strip() for p in parts if str(p or "").strip())
    if named:
        return named[:DEFAULT_TEXT_LIMIT]
    for field in fallbacks:
        value = str(values.get(field) or "").strip()
        if value:
            return value[:DEFAULT_TEXT_LIMIT]
    return None


def _defaults(fields: list[tuple[str, str, str, Any]]) -> dict[str, Any]:
    values = {field: default for _header, field, _kind, default in fields}
    values["catalog_item_id"] = None
    return values


def _parse_sheet(worksheet, fields: list[tuple[str, str, str, Any]],
                 warnings: list[str]) -> list[dict[str, Any]] | None:
    """The sheet's rows, or None when no header row could be found in it."""
    sheet_name = worksheet.title.strip() or "sheet"
    read = _cell_reader(worksheet)
    max_row = worksheet.max_row or 0
    max_col = worksheet.max_column or 0
    known = {_norm(field[0]) for field in fields}

    header_row = _find_header_row(read, max_row, max_col, known)
    if header_row is None:
        warnings.append(f"{sheet_name}: no header row found in the first "
                        f"{HEADER_SCAN_ROWS} rows, sheet skipped")
        return None

    columns = _map_columns(read, header_row, max_col, fields, sheet_name, warnings)
    if not columns:
        warnings.append(f"{sheet_name}: no known columns found, sheet skipped")
        return None

    blank = _defaults(fields)
    rows = []
    for row in range(header_row + 1, max_row + 1):
        parsed = _parse_row(read, row, columns, sheet_name, warnings)
        if parsed is not None:
            rows.append({**blank, **parsed})
    return rows


def _find_sheets(workbook, wanted: str) -> list:
    aliases = SHEET_ALIASES[wanted]
    return [workbook[title] for title in workbook.sheetnames if _norm(title) in aliases]


def _read_register(sheets: list, label: str, fields: list[tuple[str, str, str, Any]],
                   warnings: list[str], sheets_found: list[str]) -> list[dict] | None:
    """Rows of the first matching sheet; None when no matching sheet had a header."""
    if not sheets:
        warnings.append(f"No {label} sheet found")
        return []
    if len(sheets) > 1:
        titles = ", ".join(f"'{sheet.title}'" for sheet in sheets)
        warnings.append(f"Several sheets look like the {label} register ({titles}); "
                        f"only '{sheets[0].title}' was read")
    sheets_found.append(sheets[0].title)
    return _parse_sheet(sheets[0], fields, warnings)


def parse_workbook(data: bytes) -> dict[str, Any]:
    """Read the Assets/Licenses registers out of an uploaded workbook."""
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:  # openpyxl raises a zoo of types for broken files
        raise ValueError(f"could not read workbook: {exc}") from exc

    asset_sheets = _find_sheets(workbook, ASSET_SHEET)
    license_sheets = _find_sheets(workbook, LICENSE_SHEET)
    if not asset_sheets and not license_sheets:
        raise ValueError("no Assets or Licenses sheet found")

    warnings: list[str] = []
    sheets_found: list[str] = []
    assets = _read_register(asset_sheets, ASSET_SHEET, ASSET_FIELDS, warnings, sheets_found)
    licenses = _read_register(license_sheets, LICENSE_SHEET, LICENSE_FIELDS, warnings,
                              sheets_found)
    if assets is None and licenses is None or (assets is None and not license_sheets) \
            or (licenses is None and not asset_sheets):
        raise ValueError(
            "no header row found: the Assets/Licenses sheet must carry the template's "
            f"column headers within its first {HEADER_SCAN_ROWS} rows"
        )
    assets = assets or []
    licenses = licenses or []

    if len(warnings) > MAX_WARNINGS:
        hidden = len(warnings) - MAX_WARNINGS
        warnings = warnings[:MAX_WARNINGS] + [f"... and {hidden} more warnings"]

    return {
        "assets": assets,
        "licenses": licenses,
        "warnings": warnings,
        "sheets_found": sheets_found,
    }
